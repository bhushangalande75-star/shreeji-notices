from flask import Flask, render_template, request, make_response, Response, stream_with_context, session, redirect, url_for, jsonify
from flask_wtf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import pandas as pd
import os, io, zipfile, json, uuid, tempfile, glob, subprocess, sys, hmac, hashlib as _hashlib
from datetime import date, datetime, timedelta
from notice_generator import generate_notice
from notice_generator_ai import build_ai_notice_docx, build_mom_docx
from vector_kb import process_document, embed_query, extract_text
from notice_generator_2nd import generate_notice_2nd
from notice_generator_3rd import generate_notice_3rd
import requests as http_requests
import base64
from pypdf import PdfWriter, PdfReader
from database import (save_batch, get_batches, get_batch_notices, update_payment,
                      save_kb_chunks, get_kb_documents, vector_search,
                      delete_kb_document, get_kb_chunk_count,
                      get_eligible_for_2nd, get_paid_members, delete_batch,
                      get_society_by_username, get_all_societies, create_society,
                      delete_society, get_society_stats,
                      upsert_members, get_members, get_member_by_flat, delete_all_members,
                      change_society_password, reset_society_password,
                      get_stats_by_notice_type, get_unique_member_stats,
                      get_batches_by_type, create_monthly_bill, get_bills_for_society,
                      get_all_bills, update_bill_status, delete_bill, get_bill_by_id,
                      get_society_by_portal_code, set_portal_code,
                      get_member_login, upsert_member_login, update_member_pin,
                      touch_member_login, reset_member_pin,
                      get_member_notices, get_member_announcements,
                      create_announcement, delete_announcement,
                      check_password, hash_password,
                      get_society_pin_format,
                      upsert_knowledge, get_knowledge, get_member_outstanding,
                      create_ticket, get_member_tickets, get_all_tickets,
                      update_ticket_status, get_ticket_by_id,
                      log_audit, get_audit_log,
                      create_payment_order, confirm_payment, get_payment_history,
                      record_consent, has_consent, delete_member_data,
                      get_analytics)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
_secret = os.environ.get("SECRET_KEY")
if not _secret:
    raise RuntimeError(
        "SECRET_KEY environment variable is not set. "
        "Generate one with: python -c \"import secrets; print(secrets.token_hex(32))\""
    )
app.secret_key = _secret

# ── CSRF protection ────────────────────────────────────────────────────────────
csrf = CSRFProtect(app)
# Exempt JSON API routes and webhooks from CSRF (they use other auth)
CSRF_EXEMPT_PREFIXES = ("/portal/chat", "/portal/tickets/create",
                        "/whatsapp/", "/razorpay/webhook",
                        "/session/ping", "/knowledge/", "/society/",
                        "/tickets/update/", "/admin/", "/tracker/update",
                        "/tracker/delete", "/members/upload", "/members/delete",
                        "/portal/announcements/create", "/portal/announcements/delete",
                        "/ai-notices/generate", "/ai-notices/download",
                        "/ai-notices/generate-notice", "/ai-notices/generate-mom",
                        "/ai-notices/generate-committee", "/ai-notices/generate-noc")

@app.before_request
def maybe_exempt_csrf():
    path = request.path
    if request.is_json or any(path.startswith(p) for p in CSRF_EXEMPT_PREFIXES):
        setattr(request, '_csrf_token_used', True)

# ── Rate limiting ──────────────────────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://"
)

# ── External services ──────────────────────────────────────────────────────────
RAZORPAY_KEY_ID     = os.environ.get("RAZORPAY_KEY_ID", "")
RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "")
SENDGRID_API_KEY    = os.environ.get("SENDGRID_API_KEY", "")
SENDGRID_FROM_EMAIL = os.environ.get("SENDGRID_FROM_EMAIL", "noreply@societynotice.app")

# ── Session timeout: 10 minutes of inactivity ──────────────────────────────────
INACTIVITY_TIMEOUT = timedelta(minutes=10)
app.config["PERMANENT_SESSION_LIFETIME"] = INACTIVITY_TIMEOUT

@app.before_request
def check_session_timeout():
    """Log out any user who has been inactive for more than 10 minutes."""
    # Skip static files and the login/logout routes themselves
    if request.endpoint in ("login", "logout", "static", None):
        return

    # Only enforce for logged-in users
    if not session.get("society_id") and not session.get("is_admin"):
        return

    last_activity = session.get("_last_activity")
    now           = datetime.now(tz=None)

    if last_activity:
        # last_activity is stored as ISO string to survive session serialisation
        try:
            last_dt = datetime.fromisoformat(last_activity)
        except Exception:
            last_dt = None

        if last_dt and (now - last_dt) > INACTIVITY_TIMEOUT:
            session.clear()
            # For JSON/API endpoints return 401 so JS can redirect
            if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
                from flask import abort
                abort(401)
            return redirect(url_for("login"))

    # Refresh timestamp on every request
    session["_last_activity"] = now.isoformat()
    session.permanent = True      # honour PERMANENT_SESSION_LIFETIME

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin@2026")

TEMP_DIR = os.path.join(tempfile.gettempdir(), "society_notices")
os.makedirs(TEMP_DIR, exist_ok=True)

def get_libreoffice_path():
    paths = {
        "win32":  [r"C:\Program Files\LibreOffice\program\soffice.exe",
                   r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"],
        "darwin": ["/Applications/LibreOffice.app/Contents/MacOS/soffice"],
    }.get(sys.platform, ["/usr/bin/libreoffice", "/usr/bin/soffice"])
    return next((p for p in paths if os.path.exists(p)), None)

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("society_id") and not session.get("is_admin"):
            # For AJAX/fetch requests return JSON 401 instead of HTML redirect
            # so the frontend can show a proper error instead of crashing on JSON parse
            if (request.is_json
                    or request.headers.get("X-Requested-With") == "XMLHttpRequest"
                    or request.method in ("POST", "PUT", "PATCH", "DELETE")
                    and request.content_type and "multipart" in request.content_type):
                return jsonify({"success": False, "error": "Session expired. Please log in again."}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def society_required(f):
    """Allows only logged-in society users. Blocks admin."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("society_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── Auth ───────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
@limiter.limit("10 per minute")
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        ip = request.remote_addr

        # Admin login
        if username == "admin" and password == ADMIN_PASSWORD:
            session["is_admin"]   = True
            session["society_id"] = None
            session["society_name"] = "Admin"
            log_audit(None, "admin", "LOGIN", "Admin login", ip)
            return redirect(url_for("admin_dashboard"))

        # Society login
        society = get_society_by_username(username)
        if society and check_password(password, society["password"]):
            stored = society["password"]
            if not (stored.startswith("$2b$") or stored.startswith("$2a$")):
                change_society_password(society["id"], password)
            session["society_id"]   = society["id"]
            session["society_name"] = society["name"]
            session["is_admin"]     = False
            log_audit(society["id"], username, "LOGIN", "Society login", ip)
            return redirect(url_for("index"))

        log_audit(None, username, "LOGIN_FAILED", f"Failed login from {ip}", ip)
        error = "❌ Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/session/ping", methods=["POST"])
@login_required
def session_ping():
    """Lightweight endpoint called by the client to reset the inactivity timer."""
    session["_last_activity"] = datetime.now().isoformat()
    return jsonify({"ok": True})

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Admin ──────────────────────────────────────────────────────
@app.route("/admin")
@admin_required
def admin_dashboard():
    societies = get_all_societies()
    # Attach member counts to each society
    for s in societies:
        s["member_count"] = len(get_members(s["id"]))
    return render_template("admin.html", societies=societies, society_name=session.get("society_name", "Admin"))

@app.route("/admin/create_society", methods=["POST"])
@admin_required
def admin_create_society():
    sid = create_society(
        request.form.get("name"),
        request.form.get("address"),
        request.form.get("username"),
        request.form.get("password"),
        request.form.get("regd_no"),
        request.form.get("default_pin_format", "no_hyphen"),
    )
    # Optionally process member Excel if uploaded with the form
    member_file = request.files.get("member_excel")
    if member_file and member_file.filename:
        try:
            _process_member_excel(member_file, sid)
        except Exception as e:
            print(f"[WARN] Member Excel upload failed during society creation: {e}")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/delete_society/<int:society_id>", methods=["POST"])
@admin_required
def admin_delete_society(society_id):
    delete_society(society_id)
    return jsonify({"success": True})

# ── Main App ───────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    sid = session.get("society_id")
    stats = get_unique_member_stats(sid) if sid else {}
    bills = get_bills_for_society(sid) if sid else []
    return render_template("index.html", society_name=session["society_name"], stats=stats, bills=bills)

@app.route("/tracker")
@login_required
def tracker():
    society_id  = session.get("society_id")
    notice_type = request.args.get("notice_type", "")
    batches     = get_batches_by_type(society_id, notice_type or None) if society_id else []
    stats       = get_stats_by_notice_type(society_id, notice_type or None) if society_id else {}
    return render_template("tracker.html", batches=batches, society_name=session["society_name"],
                           stats=stats, selected_type=notice_type)

@app.route("/tracker/batch/<int:batch_id>")
@login_required
def batch_detail(batch_id):
    notices = get_batch_notices(batch_id)
    batches = get_batches(session["society_id"])
    batch   = next((b for b in batches if b["id"] == batch_id), None)
    paid    = [n for n in notices if n["payment_status"] == "Paid"]
    pending = [n for n in notices if n["payment_status"] == "Pending"]
    return render_template("batch_detail.html", batch=batch, notices=notices,
                           paid=paid, pending=pending, society_name=session["society_name"])

@app.route("/tracker/update_payment", methods=["POST"])
@login_required
def update_payment_route():
    data = request.json
    update_payment(data["notice_id"], data["status"], data.get("payment_date", ""),
                   data.get("payment_amount", 0), data.get("remark", ""))
    return jsonify({"success": True})

@app.route("/tracker/delete_batch/<int:batch_id>", methods=["POST"])
@login_required
def delete_batch_route(batch_id):
    delete_batch(batch_id)
    return jsonify({"success": True})

@app.route("/tracker/export/<int:batch_id>/<report_type>")
@login_required
def export_report(batch_id, report_type):
    batches = get_batches(session["society_id"])
    batch   = next((b for b in batches if b["id"] == batch_id), None)
    members = get_eligible_for_2nd(batch_id) if report_type == "eligible" else get_paid_members(batch_id)
    title   = "Eligible_for_2nd_Notice" if report_type == "eligible" else "Paid_Members"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="C00000")
    headers  = ["Flat No", "Ref No", "Member Name", "Amount (Rs.)", "Status", "Payment Date", "Amount Paid", "Remark"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for row, m in enumerate(members, 2):
        ws.cell(row=row, column=1, value=m["flat_no"])
        ws.cell(row=row, column=2, value=m["ref_no"])
        ws.cell(row=row, column=3, value=m["member_name"])
        ws.cell(row=row, column=4, value=m["amount"])
        ws.cell(row=row, column=5, value=m["payment_status"])
        ws.cell(row=row, column=6, value=m.get("payment_date", ""))
        ws.cell(row=row, column=7, value=m.get("payment_amount", ""))
        ws.cell(row=row, column=8, value=m.get("payment_remark", ""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={title}_Batch{batch_id}.xlsx"
    return response

# ── Generate ───────────────────────────────────────────────────
@app.route("/generate", methods=["POST"])
@login_required
def generate():
    if "excel" not in request.files:
        return make_response(json.dumps({"error": "No file uploaded"}), 400)
    file = request.files["excel"]
    if file.filename == "":
        return make_response(json.dumps({"error": "No file selected"}), 400)
    try:
        df   = pd.read_excel(file, header=None)
        data = df.iloc[1:].reset_index(drop=True)
    except Exception as e:
        return make_response(json.dumps({"error": f"Could not read Excel: {str(e)}"}), 400)

    total              = len(data)
    notice_type        = request.form.get("notice_type", "1st")
    batch_name         = request.form.get("batch_name", f"Batch {date.today().strftime('%b %Y')}")
    maintenance_period = request.form.get("maintenance_period", "March 2026")
    due_date           = request.form.get("due_date", "31st March 2026")
    subject            = request.form.get("subject", "Sub: Notice for Recovery of Due Maintenance.")
    raw_date           = request.form.get("issued_date", date.today().strftime("%d-%m-%Y"))
    try:
        issued_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        issued_date = raw_date

    society_id = session["society_id"]
    sess_id    = str(uuid.uuid4())
    sess_dir   = os.path.join(TEMP_DIR, sess_id)
    os.makedirs(sess_dir, exist_ok=True)
    soffice    = get_libreoffice_path()

    def stream():
        docx_files  = []
        members_log = []
        count = 0

        yield f"data: {json.dumps({'type':'start','total':total})}\n\n"

        for i, row in data.iterrows():
            try:
                flat_no     = str(row[2]).strip()
                ref_no      = str(row[4]).strip()
                name        = str(row[5]).strip()
                amount      = int(row[7])
                prev_ref_no = str(row[8]).strip() if notice_type == "2nd" and len(row) > 8 else ""
                prev_ref_no_1st = str(row[8]).strip() if notice_type == "3rd" and len(row) > 8 else ""
                prev_ref_no_2nd = str(row[9]).strip() if notice_type == "3rd" and len(row) > 9 else ""

                if notice_type == "3rd":
                    doc_bytes = generate_notice_3rd(flat_no, ref_no, name, amount, prev_ref_no_1st, prev_ref_no_2nd, issued_date, due_date, maintenance_period, subject)
                elif notice_type == "2nd":
                    doc_bytes = generate_notice_2nd(flat_no, ref_no, name, amount, prev_ref_no, issued_date, due_date, maintenance_period, subject)
                else:
                    doc_bytes = generate_notice(flat_no, ref_no, name, amount, due_date, maintenance_period, subject, issued_date)

                filename = f"Notice_{ref_no.replace('/', '-')}_{flat_no}.docx"
                docx_files.append((filename, doc_bytes))
                members_log.append({"flat_no": flat_no, "ref_no": ref_no, "name": name, "amount": amount, "prev_ref_no": prev_ref_no})
                with open(os.path.join(sess_dir, filename), "wb") as f:
                    f.write(doc_bytes)
                count += 1
                yield f"data: {json.dumps({'type':'progress','count':count,'total':total,'name':name})}\n\n"
            except Exception as e:
                print(f"Row error: {e}")

        if count == 0:
            yield f"data: {json.dumps({'type':'failed','msg':'No notices generated. Check Excel format.'})}\n\n"
            return

        save_batch(batch_name, notice_type, issued_date, members_log, society_id)

        yield f"data: {json.dumps({'type':'status','msg':'Creating ZIP file...'})}\n\n"
        zip_path = os.path.join(sess_dir, "notices.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, fbytes in docx_files:
                zf.writestr(fname, fbytes)

        pdf_pages = 0
        if soffice:
            yield f"data: {json.dumps({'type':'status','msg':'Converting to PDF...'})}\n\n"
            pdf_dir   = os.path.join(sess_dir, "pdf")
            os.makedirs(pdf_dir, exist_ok=True)
            docx_list = sorted(glob.glob(os.path.join(sess_dir, "*.docx")))
            subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir", pdf_dir] + docx_list, capture_output=True)
            yield f"data: {json.dumps({'type':'status','msg':'Merging PDF pages...'})}\n\n"
            writer = PdfWriter()
            for pf in sorted(glob.glob(os.path.join(pdf_dir, "*.pdf"))):
                for page in PdfReader(pf).pages:
                    writer.add_page(page)
            pdf_path = os.path.join(sess_dir, "notices.pdf")
            with open(pdf_path, "wb") as f:
                writer.write(f)
            pdf_pages = len(writer.pages)

        # Build member list for WhatsApp preview (flat_no → name)
        wa_members = [{"flat_no": m["flat_no"], "name": m["name"]} for m in members_log]
        yield f"data: {json.dumps({'type':'done','sess_id':sess_id,'count':count,'pages':pdf_pages,'has_pdf':soffice is not None,'wa_members':wa_members})}\n\n"

    return Response(stream_with_context(stream()), mimetype="text/event-stream")

@app.route("/download/<sess_id>/<filetype>")
@login_required
def download(sess_id, filetype):
    sess_dir = os.path.join(TEMP_DIR, sess_id)
    paths = {
        "zip": (os.path.join(sess_dir, "notices.zip"), "application/zip",  "Maintenance_Notices.zip"),
        "pdf": (os.path.join(sess_dir, "notices.pdf"), "application/pdf",  "Maintenance_Notices_All.pdf"),
    }
    if filetype not in paths: return "Invalid", 400
    path, mime, name = paths[filetype]
    if not os.path.exists(path): return "File not found", 404
    response = make_response(open(path, "rb").read())
    response.headers["Content-Type"]        = mime
    response.headers["Content-Disposition"] = f"attachment; filename={name}"
    return response

@app.route("/ai-notices")
@login_required
def ai_notices():
    return render_template("ai_notices.html",
                           society_name=session["society_name"])

GROQ_API_KEY = os.environ.get("NOTICE_API_KEY", "")
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL   = "llama-3.3-70b-versatile"

def call_groq(system_prompt, user_content):
    """Call Groq API (OpenAI-compatible). user_content can be str or list (for vision)."""
    if not GROQ_API_KEY:
        raise ValueError(
            "Groq API key is not set. "
            "Add NOTICE_API_KEY in your Render environment variables."
        )

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type":  "application/json",
    }

    if isinstance(user_content, str):
        model = GROQ_MODEL
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_content},
        ]
    else:
        # Vision input — use Groq vision model
        model = "meta-llama/llama-4-scout-17b-16e-instruct"
        content_parts = []
        for item in user_content:
            if item.get("type") == "text":
                content_parts.append({"type": "text", "text": item["text"]})
            elif item.get("type") == "image":
                src = item["source"]
                data_url = f"data:{src['media_type']};base64,{src['data']}"
                content_parts.append({
                    "type":      "image_url",
                    "image_url": {"url": data_url},
                })
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": content_parts},
        ]

    payload = {
        "model":       model,
        "messages":    messages,
        "max_tokens":  2048,
        "temperature": 0.4,
    }

    resp = http_requests.post(GROQ_API_URL, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    return data["choices"][0]["message"]["content"]


@app.route("/ai-notices/generate-notice", methods=["POST"])
@csrf.exempt
@login_required
def ai_generate_notice():
    """Generate a notice in the user's chosen language (English / Marathi / Hindi)."""
    try:
        notice_type   = request.form.get("notice_type", "General Notice")
        flat_no       = request.form.get("flat_no", "").strip()
        member_name   = request.form.get("member_name", "").strip()
        ref_no        = request.form.get("ref_no", "").strip()
        issued_date   = request.form.get("issued_date", date.today().strftime("%d-%m-%Y"))
        description   = request.form.get("description", "").strip()
        language      = request.form.get("language", "English").strip()
        society_name  = session.get("society_name", "Shreeji Iconic CHS Ltd.")

        try:
            issued_date = datetime.strptime(issued_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass

        # Language-specific config
        lang_cfg = {
            "English": {
                "system": (
                    "You are the official notice writer for a Co-operative Housing Society in Maharashtra, India. "
                    "Write the notice in formal, firm, and legally appropriate English. "
                    "Write ONLY the body paragraphs — no salutation, no subject line, no signature. "
                    "Each paragraph on a new line. "
                    "Reference the Maharashtra Co-operative Societies Act 1960 and society bye-laws where relevant."
                ),
                "user": (
                    f"Write a formal notice in English for {society_name} for the following situation:\n\n"
                    f"Notice Type: {notice_type}\n"
                    f"Member Name: {member_name}\n"
                    f"Flat No: {flat_no}\n"
                    f"Issue Description: {description}\n\n"
                    "Write 3-4 firm but respectful paragraphs covering: what the issue is, "
                    "how it violates society rules/bye-laws, a demand to stop/rectify immediately, "
                    "and consequences if not complied with."
                ),
                "subject_system": "You are the secretary of a Co-operative Housing Society.",
                "subject_user": (
                    f"Write a one-line formal English subject line for a notice regarding '{notice_type}'. "
                    "Output ONLY the subject text — no 'Sub:', 'Subject:' or any prefix. "
                    "Example: No Parking Violation — Immediate Compliance Required."
                ),
                "sub_label": "Sub:",
            },
            "Marathi": {
                "system": (
                    "तुम्ही महाराष्ट्रातील एका सहकारी गृहनिर्माण संस्थेचे अधिकृत नोटीस लेखक आहात. "
                    "नोटीस मराठीत लिहा — औपचारिक, ठाम आणि कायदेशीरदृष्ट्या योग्य भाषेत. "
                    "फक्त नोटीसचे मुख्य परिच्छेद लिहा — अभिवादन, विषय ओळ किंवा स्वाक्षरी नको. "
                    "प्रत्येक परिच्छेद नवीन ओळीवर लिहा. "
                    "महाराष्ट्र सहकारी संस्था अधिनियम १९६० आणि संस्थेच्या उपविधींचा संदर्भ द्या."
                ),
                "user": (
                    f"{society_name} येथील खालील परिस्थितीसाठी मराठीत औपचारिक नोटीस लिहा:\n\n"
                    f"नोटीस प्रकार: {notice_type}\n"
                    f"सदस्याचे नाव: {member_name}\n"
                    f"फ्लॅट क्र.: {flat_no}\n"
                    f"समस्येचे वर्णन: {description}\n\n"
                    "३-४ ठाम पण सभ्य परिच्छेद लिहा. "
                    "समाविष्ट करा: समस्या काय आहे, ती संस्थेच्या नियमांचे/उपविधींचे उल्लंघन कसे करते, "
                    "तात्काळ थांबण्याची/दुरुस्त करण्याची मागणी, आणि पालन न केल्यास होणारे परिणाम."
                ),
                "subject_system": "तुम्ही एका सहकारी गृहनिर्माण संस्थेचे सचिव आहात.",
                "subject_user": (
                    f"'{notice_type}' या विषयावरील नोटीससाठी एक ओळीचा औपचारिक मराठी विषय लिहा. "
                    "फक्त विषय ओळ लिहा — 'विषय:', 'Sub:' किंवा इतर कोणताही उपसर्ग न लिहिता. "
                    "उदा: पार्किंगच्या नियमांचे उल्लंघन — तात्काळ अनुपालन आवश्यक."
                ),
                "sub_label": "विषय:",
            },
            "Hindi": {
                "system": (
                    "आप महाराष्ट्र की एक सहकारी आवास संस्था के आधिकारिक नोटिस लेखक हैं. "
                    "नोटिस हिंदी में लिखें — औपचारिक, दृढ़ और कानूनी रूप से उचित भाषा में. "
                    "केवल नोटिस के मुख्य अनुच्छेद लिखें — अभिवादन, विषय पंक्ति या हस्ताक्षर नहीं. "
                    "प्रत्येक अनुच्छेद नई पंक्ति पर लिखें. "
                    "महाराष्ट्र सहकारी संस्था अधिनियम 1960 और संस्था के उपनियमों का संदर्भ दें."
                ),
                "user": (
                    f"{society_name} के लिए निम्नलिखित स्थिति हेतु हिंदी में औपचारिक नोटिस लिखें:\n\n"
                    f"नोटिस प्रकार: {notice_type}\n"
                    f"सदस्य का नाम: {member_name}\n"
                    f"फ्लैट नं.: {flat_no}\n"
                    f"समस्या का विवरण: {description}\n\n"
                    "3-4 दृढ़ लेकिन शिष्ट अनुच्छेद लिखें जिनमें शामिल हों: समस्या क्या है, "
                    "यह संस्था के नियमों/उपनियमों का उल्लंघन कैसे करती है, "
                    "तत्काल रोकने/सुधारने की मांग, और पालन न करने पर परिणाम."
                ),
                "subject_system": "आप एक सहकारी आवास संस्था के सचिव हैं.",
                "subject_user": (
                    f"'{notice_type}' विषय पर नोटिस के लिए एक पंक्ति का औपचारिक हिंदी विषय लिखें. "
                    "केवल विषय पंक्ति लिखें — 'विषय:', 'Sub:' या कोई उपसर्ग नहीं. "
                    "उदा: पार्किंग नियमों का उल्लंघन — तत्काल अनुपालन आवश्यक."
                ),
                "sub_label": "विषय:",
            },
        }

        cfg = lang_cfg.get(language, lang_cfg["English"])
        print(f"[AI-NOTICE] language={language!r}  sub_label={cfg['sub_label']!r}")

        ai_text     = call_groq(cfg["system"], cfg["user"])
        raw_subject = call_groq(cfg["subject_system"], cfg["subject_user"]).strip().strip('"').strip("'").strip()
        # Strip any prefix the model may have added despite instructions
        for prefix in ("Sub:", "Subject:", "विषय:", "विषय :", "Vishay:"):
            if raw_subject.lower().startswith(prefix.lower()):
                raw_subject = raw_subject[len(prefix):].strip()
                break
        subject = f"{cfg['sub_label']} {raw_subject}"

        docx_bytes = build_ai_notice_docx(ref_no, flat_no, member_name, issued_date, subject, ai_text)

        sess_id  = str(uuid.uuid4())
        sess_dir = os.path.join(TEMP_DIR, sess_id)
        os.makedirs(sess_dir, exist_ok=True)
        safe_notice = notice_type.replace(' ', '_').replace('/', '-').replace('\\', '-').replace(':', '').replace('(', '').replace(')', '').replace('*', '').replace('?', '').replace('"', '').replace('<', '').replace('>', '').replace('|', '')
        fname = f"Notice_{safe_notice}_{flat_no or 'General'}.docx"
        with open(os.path.join(sess_dir, fname), "wb") as f:
            f.write(docx_bytes)

        return jsonify({"success": True, "preview": ai_text, "sess_id": sess_id, "filename": fname})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ai-notices/generate-mom", methods=["POST"])
@csrf.exempt
@login_required
def ai_generate_mom():
    """Generate MOM in the user's chosen language (English / Marathi / Hindi)."""
    try:
        meeting_date = request.form.get("meeting_date", date.today().strftime("%d-%m-%Y"))
        attendees    = request.form.get("attendees", "").strip()
        language     = request.form.get("language", "English").strip()
        society_name = session.get("society_name", "Shreeji Iconic CHS Ltd.")

        try:
            meeting_date = datetime.strptime(meeting_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass

        # Language-specific prompts for MOM
        mom_lang_cfg = {
            "English": {
                "system": (
                    "You are an experienced secretary of a Co-operative Housing Society in Maharashtra, India. "
                    "You write fluent, formal Minutes of Meeting in English. "
                    "Number all decisions. Use proper legal and administrative terminology. "
                    "Sections: Members Present, Agenda, Discussion & Decisions (numbered), Action Items."
                ),
                "photo_text": (
                    f"This is a handwritten meeting notes photo for {society_name}. "
                    f"Meeting date: {meeting_date}. "
                    f"Attendees: {attendees or 'As visible in the notes'}.\n\n"
                    "Please read all the handwritten content and generate a complete, "
                    "formal Minutes of Meeting in English. "
                    "Sections: Members Present, Agenda, Discussion & Decisions (numbered), Action Items. "
                    "Output ONLY the MOM content, no preamble."
                ),
                "text_prompt": (
                    f"Generate formal Minutes of Meeting in English for {society_name}.\n"
                    f"Meeting date: {meeting_date}\nAttendees: {attendees}\n"
                    f"Meeting notes: {{raw_notes}}\n\n"
                    "Sections: Members Present, Agenda, Discussion & Decisions (numbered), Action Items. "
                    "Output ONLY the MOM content."
                ),
            },
            "Marathi": {
                "system": (
                    "तुम्ही महाराष्ट्रातील एका सहकारी गृहनिर्माण संस्थेचे अनुभवी सचिव आहात. "
                    "तुम्ही अस्खलित, औपचारिक मराठीत इतिवृत्त (Minutes of Meeting) लिहिता. "
                    "निर्णय क्रमांकित करा. योग्य मराठी कायदेशीर आणि प्रशासकीय शब्दावली वापरा. "
                    "विभाग: उपस्थित सदस्य, अजेंडा, चर्चा व निर्णय (क्रमांकित), कृती मुद्दे."
                ),
                "photo_text": (
                    f"हे {society_name} च्या बैठकीच्या हस्तलिखित नोट्सचा फोटो आहे. "
                    f"बैठकीची तारीख: {meeting_date}. "
                    f"उपस्थित सदस्य: {attendees or 'नोट्समध्ये दिसत आहे'}.\n\n"
                    "कृपया सर्व हस्तलिखित मजकूर वाचा आणि संपूर्ण, औपचारिक मराठी इतिवृत्त तयार करा. "
                    "विभाग: उपस्थित सदस्य, अजेंडा, चर्चा व निर्णय (क्रमांकित), कृती मुद्दे. "
                    "फक्त इतिवृत्त मजकूर लिहा, कोणतीही प्रस्तावना नको."
                ),
                "text_prompt": (
                    f"{society_name} साठी मराठीत औपचारिक इतिवृत्त तयार करा.\n"
                    f"बैठकीची तारीख: {meeting_date}\nउपस्थित सदस्य: {attendees}\n"
                    f"बैठकीच्या नोट्स: {{raw_notes}}\n\n"
                    "विभाग: उपस्थित सदस्य, अजेंडा, चर्चा व निर्णय (क्रमांकित), कृती मुद्दे. "
                    "फक्त इतिवृत्त मजकूर लिहा."
                ),
            },
            "Hindi": {
                "system": (
                    "आप महाराष्ट्र की एक सहकारी आवास संस्था के अनुभवी सचिव हैं. "
                    "आप धाराप्रवाह, औपचारिक हिंदी में कार्यवृत्त (Minutes of Meeting) लिखते हैं. "
                    "सभी निर्णयों को क्रमांकित करें. उचित कानूनी और प्रशासनिक शब्दावली का उपयोग करें. "
                    "खंड: उपस्थित सदस्य, एजेंडा, चर्चा और निर्णय (क्रमांकित), कार्य बिंदु."
                ),
                "photo_text": (
                    f"यह {society_name} की बैठक के हस्तलिखित नोट्स का फोटो है. "
                    f"बैठक की तारीख: {meeting_date}. "
                    f"उपस्थित सदस्य: {attendees or 'नोट्स में दिखाई दे रहे हैं'}.\n\n"
                    "कृपया सभी हस्तलिखित सामग्री पढ़ें और पूर्ण, औपचारिक हिंदी कार्यवृत्त तैयार करें. "
                    "खंड: उपस्थित सदस्य, एजेंडा, चर्चा और निर्णय (क्रमांकित), कार्य बिंदु. "
                    "केवल कार्यवृत्त सामग्री लिखें, कोई प्रस्तावना नहीं."
                ),
                "text_prompt": (
                    f"{society_name} के लिए हिंदी में औपचारिक कार्यवृत्त तैयार करें.\n"
                    f"बैठक की तारीख: {meeting_date}\nउपस्थित सदस्य: {attendees}\n"
                    f"बैठक के नोट्स: {{raw_notes}}\n\n"
                    "खंड: उपस्थित सदस्य, एजेंडा, चर्चा और निर्णय (क्रमांकित), कार्य बिंदु. "
                    "केवल कार्यवृत्त सामग्री लिखें."
                ),
            },
        }

        cfg = mom_lang_cfg.get(language, mom_lang_cfg["English"])

        if "photo" in request.files and request.files["photo"].filename:
            photo     = request.files["photo"]
            img_bytes = photo.read()
            img_b64   = base64.standard_b64encode(img_bytes).decode()
            mime      = photo.content_type or "image/jpeg"
            user_content = [
                {"type": "image", "source": {"type": "base64", "media_type": mime, "data": img_b64}},
                {"type": "text",  "text": cfg["photo_text"]},
            ]
        else:
            raw_notes    = request.form.get("raw_notes", "").strip()
            user_content = cfg["text_prompt"].format(raw_notes=raw_notes)

        mom_text   = call_groq(cfg["system"], user_content)
        docx_bytes = build_mom_docx(mom_text, meeting_date, society_name)

        sess_id  = str(uuid.uuid4())
        sess_dir = os.path.join(TEMP_DIR, sess_id)
        os.makedirs(sess_dir, exist_ok=True)
        fname = f"MOM_{meeting_date.replace('/', '-')}.docx"
        with open(os.path.join(sess_dir, fname), "wb") as f:
            f.write(docx_bytes)

        return jsonify({"success": True, "preview": mom_text, "sess_id": sess_id, "filename": fname})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ai-notices/generate-committee", methods=["POST"])
@csrf.exempt
@login_required
def ai_generate_committee():
    """Generate a formal committee meeting notice with agenda in chosen language."""
    try:
        meeting_date  = request.form.get("meeting_date", "").strip()
        meeting_time  = request.form.get("meeting_time", "").strip()
        venue         = request.form.get("venue", "").strip()
        ref_no        = request.form.get("ref_no", "").strip()
        notice_date   = request.form.get("notice_date", date.today().strftime("%d-%m-%Y"))
        meeting_type  = request.form.get("meeting_type", "Managing Committee Meeting")
        members       = request.form.get("members", "").strip()
        agenda        = request.form.get("agenda", "").strip()
        notes         = request.form.get("notes", "").strip()
        language      = request.form.get("language", "English").strip()
        society_name  = session.get("society_name", "Shreeji Iconic CHS Ltd.")

        try:
            meeting_date = datetime.strptime(meeting_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass
        try:
            notice_date = datetime.strptime(notice_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass

        # Format agenda items
        agenda_lines = [l.strip() for l in agenda.replace(",", "\n").split("\n") if l.strip()]
        agenda_numbered = "\n".join(
            f"{i+1}. {item}" if not item[0].isdigit() else item
            for i, item in enumerate(agenda_lines)
        )

        lang_cfg = {
            "English": {
                "system": (
                    "You are the official secretary of a Co-operative Housing Society in Maharashtra, India. "
                    "Write a formal committee meeting notice in English. "
                    "Include: formal salutation to committee members, meeting details, numbered agenda, "
                    "and a request to confirm attendance. "
                    "Write ONLY the body paragraphs — no letterhead, no subject line, no signature block. "
                    "Reference the Maharashtra Co-operative Societies Act 1960 where relevant."
                ),
                "user": (
                    f"Write a formal {meeting_type} notice for {society_name}.\n\n"
                    f"Meeting Date: {meeting_date}\n"
                    f"Time: {meeting_time}\n"
                    f"Venue: {venue}\n"
                    f"Committee Members: {members or 'All Committee Members'}\n\n"
                    f"Agenda:\n{agenda_numbered}\n\n"
                    f"{('Additional Notes: ' + notes) if notes else ''}\n\n"
                    "Write 2-3 opening paragraphs formally notifying members of the meeting, "
                    "then list the agenda items clearly, then close with attendance confirmation request."
                ),
                "subject_system": "You are the secretary of a Co-operative Housing Society.",
                "subject_user": (
                    f"Write a one-line formal English subject for a {meeting_type} notice. "
                    f"Meeting date: {meeting_date}. "
                    "Output ONLY the subject text, no prefix like 'Sub:' or 'Subject:'."
                ),
                "sub_label": "Sub:",
            },
            "Marathi": {
                "system": (
                    "तुम्ही महाराष्ट्रातील एका सहकारी गृहनिर्माण संस्थेचे अधिकृत सचिव आहात. "
                    "समिती बैठकीची नोटीस मराठीत लिहा — औपचारिक आणि कायदेशीरदृष्ट्या योग्य भाषेत. "
                    "फक्त मुख्य परिच्छेद लिहा — लेटरहेड, विषय ओळ किंवा स्वाक्षरी नको."
                ),
                "user": (
                    f"{society_name} च्या {meeting_type} साठी मराठीत औपचारिक नोटीस लिहा.\n\n"
                    f"बैठकीची तारीख: {meeting_date}\n"
                    f"वेळ: {meeting_time}\n"
                    f"ठिकाण: {venue}\n"
                    f"समिती सदस्य: {members or 'सर्व समिती सदस्य'}\n\n"
                    f"अजेंडा:\n{agenda_numbered}\n\n"
                    f"{('अतिरिक्त सूचना: ' + notes) if notes else ''}\n\n"
                    "सुरुवातीला बैठकीची औपचारिक सूचना द्या, नंतर अजेंडा क्रमांकित करा, "
                    "शेवटी उपस्थितीची पुष्टी करण्याची विनंती करा."
                ),
                "subject_system": "तुम्ही एका सहकारी गृहनिर्माण संस्थेचे सचिव आहात.",
                "subject_user": (
                    f"'{meeting_type}' साठी एक ओळीचा औपचारिक मराठी विषय लिहा. "
                    f"बैठकीची तारीख: {meeting_date}. "
                    "फक्त विषय ओळ लिहा — 'विषय:' किंवा इतर कोणताही उपसर्ग न लिहिता."
                ),
                "sub_label": "विषय:",
            },
            "Hindi": {
                "system": (
                    "आप महाराष्ट्र की एक सहकारी आवास संस्था के आधिकारिक सचिव हैं. "
                    "समिति बैठक की सूचना हिंदी में लिखें — औपचारिक और कानूनी रूप से उचित भाषा में. "
                    "केवल मुख्य अनुच्छेद लिखें — लेटरहेड, विषय पंक्ति या हस्ताक्षर नहीं."
                ),
                "user": (
                    f"{society_name} की {meeting_type} के लिए हिंदी में औपचारिक सूचना लिखें.\n\n"
                    f"बैठक की तारीख: {meeting_date}\n"
                    f"समय: {meeting_time}\n"
                    f"स्थान: {venue}\n"
                    f"समिति सदस्य: {members or 'सभी समिति सदस्य'}\n\n"
                    f"एजेंडा:\n{agenda_numbered}\n\n"
                    f"{('अतिरिक्त निर्देश: ' + notes) if notes else ''}\n\n"
                    "औपचारिक रूप से बैठक की सूचना दें, एजेंडा क्रमांकित करें, "
                    "और उपस्थिति की पुष्टि करने का अनुरोध करें."
                ),
                "subject_system": "आप एक सहकारी आवास संस्था के सचिव हैं.",
                "subject_user": (
                    f"'{meeting_type}' के लिए एक पंक्ति का औपचारिक हिंदी विषय लिखें. "
                    f"बैठक की तारीख: {meeting_date}. "
                    "केवल विषय पंक्ति लिखें — 'विषय:' या कोई उपसर्ग नहीं."
                ),
                "sub_label": "विषय:",
            },
        }

        cfg = lang_cfg.get(language, lang_cfg["English"])

        ai_text     = call_groq(cfg["system"], cfg["user"])
        raw_subject = call_groq(cfg["subject_system"], cfg["subject_user"]).strip().strip('"').strip("'").strip()
        for prefix in ("Sub:", "Subject:", "विषय:", "विषय :", "Vishay:"):
            if raw_subject.lower().startswith(prefix.lower()):
                raw_subject = raw_subject[len(prefix):].strip()
                break
        subject = f"{cfg['sub_label']} {raw_subject}"

        docx_bytes = build_ai_notice_docx(ref_no, "", members or "Committee Members",
                                          notice_date, subject, ai_text)

        sess_id  = str(uuid.uuid4())
        sess_dir = os.path.join(TEMP_DIR, sess_id)
        os.makedirs(sess_dir, exist_ok=True)
        safe_type = meeting_type.replace(" ", "_").replace("(", "").replace(")", "")
        fname = f"Committee_Notice_{safe_type}_{meeting_date.replace('-','')}.docx"
        with open(os.path.join(sess_dir, fname), "wb") as f:
            f.write(docx_bytes)

        return jsonify({"success": True, "preview": ai_text, "sess_id": sess_id, "filename": fname})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500



@app.route("/ai-notices/generate-noc", methods=["POST"])
@csrf.exempt
@login_required
def ai_generate_noc():
    """Generate an AI-written NOC (No Objection Certificate) in chosen language."""
    try:
        noc_type     = request.form.get("noc_type", "").strip()
        custom_type  = request.form.get("custom_type", "").strip()
        name         = request.form.get("name", "").strip()
        flat_no      = request.form.get("flat_no", "").strip()
        noc_date     = request.form.get("noc_date", date.today().strftime("%d-%m-%Y"))
        ref_no       = request.form.get("ref_no", "").strip()
        buyer        = request.form.get("buyer", "").strip()
        bank         = request.form.get("bank", "").strip()
        details      = request.form.get("details", "").strip()
        language     = request.form.get("language", "English").strip()
        society_name = session.get("society_name", "Shreeji Iconic CHS Ltd.")

        display_type = custom_type if noc_type == "Custom" else noc_type

        try:
            noc_date = datetime.strptime(noc_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass

        lang_cfg = {
            "English": {
                "system": (
                    f"You are the official secretary of {society_name}, a Co-operative Housing Society in Maharashtra, India. "
                    f"Write a formal NOC (No Objection Certificate) in English. "
                    "Write ONLY the body paragraphs — no letterhead, no subject line, no signature block. "
                    "Keep it concise, formal and legally appropriate under Maharashtra Co-operative Societies Act 1960."
                ),
                "user": (
                    f"Write a formal NOC for: {display_type}\n\n"
                    f"Member Name: {name}\nFlat No: {flat_no}\nDate: {noc_date}\n"
                    f"{'Buyer/Transferee: ' + buyer if buyer else ''}\n"
                    f"{'Bank/Lender: ' + bank if bank else ''}\n"
                    f"{'Additional Details: ' + details if details else ''}\n\n"
                    "Write 2-3 formal paragraphs: certify there is no objection, state the NOC is issued in good faith "
                    "based on society records, and mention any standard conditions/caveats if applicable."
                ),
                "subject_system": "You are the secretary of a Co-operative Housing Society in Maharashtra.",
                "subject_user": (
                    f"Write a one-line formal English subject line for a NOC regarding '{display_type}'. "
                    "Output ONLY the subject text — no 'Sub:' prefix. "
                    "Example: No Objection Certificate for Sale of Flat — Flat B01-310."
                ),
                "sub_label": "Sub:",
            },
            "Marathi": {
                "system": (
                    f"तुम्ही {society_name} या सहकारी गृहनिर्माण संस्थेचे अधिकृत सचिव आहात. "
                    "मराठीत औपचारिक ना-हरकत प्रमाणपत्र (NOC) लिहा — कायदेशीरदृष्ट्या योग्य भाषेत. "
                    "फक्त मुख्य परिच्छेद लिहा — लेटरहेड, विषय ओळ किंवा स्वाक्षरी नको."
                ),
                "user": (
                    f"'{display_type}' साठी मराठीत औपचारिक ना-हरकत प्रमाणपत्र लिहा.\n\n"
                    f"सदस्याचे नाव: {name}\nफ्लॅट क्र.: {flat_no}\nतारीख: {noc_date}\n"
                    f"{'खरेदीदार: ' + buyer if buyer else ''}\n"
                    f"{'बँक: ' + bank if bank else ''}\n"
                    f"{'तपशील: ' + details if details else ''}\n\n"
                    "२-३ औपचारिक परिच्छेद लिहा: ना-हरकत असल्याचे प्रमाणित करा, "
                    "संस्थेच्या नोंदींनुसार हे प्रमाणपत्र सदिच्छेने दिले जात असल्याचे नमूद करा."
                ),
                "subject_system": "तुम्ही एका सहकारी गृहनिर्माण संस्थेचे सचिव आहात.",
                "subject_user": (
                    f"'{display_type}' साठी एक ओळीचा औपचारिक मराठी विषय लिहा. "
                    "फक्त विषय ओळ लिहा — कोणताही उपसर्ग न लिहिता."
                ),
                "sub_label": "विषय:",
            },
            "Hindi": {
                "system": (
                    f"आप {society_name} सहकारी आवास संस्था के आधिकारिक सचिव हैं. "
                    "हिंदी में औपचारिक अनापत्ति प्रमाण पत्र (NOC) लिखें — कानूनी रूप से उचित भाषा में. "
                    "केवल मुख्य अनुच्छेद लिखें — लेटरहेड, विषय पंक्ति या हस्ताक्षर नहीं."
                ),
                "user": (
                    f"'{display_type}' के लिए हिंदी में औपचारिक अनापत्ति प्रमाण पत्र लिखें.\n\n"
                    f"सदस्य का नाम: {name}\nफ्लैट नं.: {flat_no}\nतारीख: {noc_date}\n"
                    f"{'खरीदार: ' + buyer if buyer else ''}\n"
                    f"{'बैंक: ' + bank if bank else ''}\n"
                    f"{'विवरण: ' + details if details else ''}\n\n"
                    "2-3 औपचारिक अनुच्छेद लिखें: अनापत्ति प्रमाणित करें, "
                    "संस्था के रिकॉर्ड के आधार पर सद्भावना से जारी किया गया है यह उल्लेख करें."
                ),
                "subject_system": "आप एक सहकारी आवास संस्था के सचिव हैं.",
                "subject_user": (
                    f"'{display_type}' के लिए एक पंक्ति का औपचारिक हिंदी विषय लिखें. "
                    "केवल विषय पंक्ति लिखें — कोई उपसर्ग नहीं."
                ),
                "sub_label": "विषय:",
            },
        }

        cfg = lang_cfg.get(language, lang_cfg["English"])

        ai_text     = call_groq(cfg["system"], cfg["user"])
        raw_subject = call_groq(cfg["subject_system"], cfg["subject_user"]).strip().strip('"').strip("'").strip()
        for prefix in ("Sub:", "Subject:", "विषय:", "विषय :", "Vishay:"):
            if raw_subject.lower().startswith(prefix.lower()):
                raw_subject = raw_subject[len(prefix):].strip()
                break
        subject = f"{cfg['sub_label']} {raw_subject}"

        docx_bytes = build_ai_notice_docx(ref_no, flat_no, name, noc_date, subject, ai_text)

        sess_id  = str(uuid.uuid4())
        sess_dir = os.path.join(TEMP_DIR, sess_id)
        os.makedirs(sess_dir, exist_ok=True)
        safe_type = display_type.replace(" ", "_").replace("/", "-").replace(":", "")
        fname = f"NOC_{safe_type}_{flat_no or 'General'}_{noc_date.replace('-', '')}.docx"
        with open(os.path.join(sess_dir, fname), "wb") as f:
            f.write(docx_bytes)

        return jsonify({"success": True, "preview": ai_text, "sess_id": sess_id, "filename": fname})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ai-notices/download/<sess_id>/<filename>")
@login_required
def ai_download(sess_id, filename):
    # Fix 4: sanitize filename — strips any path components (e.g. ../../etc/passwd)
    safe_filename = os.path.basename(filename)
    sess_dir = os.path.join(TEMP_DIR, sess_id)
    path = os.path.join(sess_dir, safe_filename)
    if not os.path.exists(path):
        return "File not found", 404
    with open(path, "rb") as f:
        data = f.read()
    response = make_response(data)
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    response.headers["Content-Disposition"] = f"attachment; filename={safe_filename}"
    return response


# ── AI Notices & MOM ───────────────────────────────────────────
@app.route("/whatsapp/batch-preview", methods=["POST"])
@society_required
def whatsapp_batch_preview():
    """
    Given a sess_id and list of {flat_no, name}, look up each member's phone
    from the directory. Returns a preview list for admin confirmation.
    """
    data       = request.json
    sess_id    = data.get("sess_id", "")
    wa_members = data.get("wa_members", [])  # [{flat_no, name}, ...]
    society_id = session["society_id"]

    results = []
    for m in wa_members:
        flat_combo = m["flat_no"].strip().upper()
        member     = get_member_by_flat(society_id, flat_combo)
        results.append({
            "flat_no" : m["flat_no"],
            "name"    : m["name"],
            "phone"   : member["phone"] if member else None,
            "found"   : member is not None,
        })

    found   = sum(1 for r in results if r["found"])
    missing = sum(1 for r in results if not r["found"])

    return jsonify({
        "success": True,
        "members": results,
        "found"  : found,
        "missing": missing,
        "sess_id": sess_id,
    })


@app.route("/whatsapp/send-batch", methods=["POST"])
@society_required
def whatsapp_send_batch():
    """
    Send individual PDF notices to each matched member via WhatsApp.
    Streams back SSE-style JSON lines so the UI can show progress.
    """
    data         = request.json
    sess_id      = data.get("sess_id", "")
    members      = data.get("members", [])  # [{flat_no, name, phone}, ...]
    society_name = session.get("society_name", "Society")
    society_id   = session["society_id"]
    sess_dir     = os.path.join(TEMP_DIR, sess_id)

    def stream():
        sent = 0; failed = 0
        for m in members:
            if not m.get("phone"):
                yield f"data: {json.dumps({'flat_no':m['flat_no'],'name':m['name'],'status':'skipped','reason':'No phone found'})}\n\n"
                failed += 1
                continue

            # Find the docx for this flat_no
            flat_safe  = m["flat_no"].replace("/", "-").replace(" ", "_")
            # Search for docx matching this flat
            candidates = glob.glob(os.path.join(sess_dir, f"*_{m['flat_no']}*.docx"))
            if not candidates:
                candidates = glob.glob(os.path.join(sess_dir, f"*{flat_safe}*.docx"))
            if not candidates:
                yield f"data: {json.dumps({'flat_no':m['flat_no'],'name':m['name'],'status':'skipped','reason':'Notice file not found'})}\n\n"
                failed += 1
                continue

            docx_path = candidates[0]
            pdf_path  = _docx_to_pdf(docx_path)
            if not pdf_path:
                yield f"data: {json.dumps({'flat_no':m['flat_no'],'name':m['name'],'status':'failed','reason':'PDF conversion failed'})}\n\n"
                failed += 1
                continue

            pdf_filename = os.path.basename(pdf_path)
            caption      = (f"Dear {m['name']},\n\n"
                            f"Please find attached the official notice from {society_name}.\n\n"
                            f"Regards,\n{society_name}")

            success, msg = _send_whatsapp_document(m["phone"], pdf_path, caption, pdf_filename)
            if success:
                sent += 1
                yield f"data: {json.dumps({'flat_no':m['flat_no'],'name':m['name'],'status':'sent','phone':m['phone']})}\n\n"
            else:
                failed += 1
                yield f"data: {json.dumps({'flat_no':m['flat_no'],'name':m['name'],'status':'failed','reason':msg})}\n\n"

        yield f"data: {json.dumps({'type':'complete','sent':sent,'failed':failed})}\n\n"

    return Response(stream_with_context(stream()), mimetype="text/event-stream")


def _process_member_excel(file_obj, society_id):
    """
    Parse a member directory Excel/CSV and upsert into society_members.
    Returns count of members processed.
    """
    df = pd.read_excel(file_obj)
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    col_map = {
        "building_no": ["building_no", "building", "bldg", "bldg_no"],
        "flat_no"    : ["flat_no", "flat", "unit", "unit_no"],
        "flat_combo" : ["flat_combo", "combo", "flat_combination", "building_flat"],
        "name"       : ["name", "member_name", "owner_name", "resident_name"],
        "phone"      : ["phone", "mobile", "phone_no", "mobile_no", "contact"],
        "email"      : ["email", "email_id", "email_address"],
    }

    def find_col(key):
        for alias in col_map[key]:
            if alias in df.columns:
                return alias
        return None

    building_col = find_col("building_no")
    flat_col     = find_col("flat_no")
    combo_col    = find_col("flat_combo")
    name_col     = find_col("name")
    phone_col    = find_col("phone")
    email_col    = find_col("email")

    if not name_col or not phone_col:
        raise ValueError("Excel must have Name and Phone columns")

    members = []
    for _, row in df.iterrows():
        name  = str(row.get(name_col, "")).strip()
        # Clean phone — remove spaces and dots but NOT hyphens (needed for flat combos)
        phone_raw = str(row.get(phone_col, "")).strip()
        phone = phone_raw.replace(" ", "").replace(".", "").replace("+", "")
        # Remove leading country code if 12 digits starting with 91
        if len(phone) == 12 and phone.startswith("91"):
            phone = phone[2:]
        if not name or not phone or name.lower() == "nan" or phone.lower() == "nan":
            continue
        building = str(row.get(building_col, "")).strip() if building_col else ""
        flat     = str(row.get(flat_col,     "")).strip() if flat_col     else ""
        email    = str(row.get(email_col,    "")).strip() if email_col    else ""
        email    = "" if email.lower() in ("nan", "") else email

        # Build flat_combo: prefer explicit combo col, else B01-310 style
        if combo_col:
            combo = str(row.get(combo_col, "")).strip()
        elif building and flat:
            # e.g. building=B01, flat=310  →  B01-310
            combo = f"{building}-{flat}"
        elif flat:
            # flat already contains full combo like B01-310 or B01001
            combo = flat
        else:
            combo = building

        # Normalise: strip whitespace, uppercase
        combo = combo.strip().upper()
        if not combo or combo == "NAN":
            continue

        # If building not set but combo looks like B01-310, extract building from combo
        if not building and "-" in combo:
            parts = combo.split("-", 1)
            building = parts[0]
            flat     = parts[1] if not flat else flat

        members.append({
            "building_no": building.upper() if building else "",
            "flat_no"    : flat.upper() if flat else combo,
            "flat_combo" : combo,
            "name"       : name,
            "phone"      : phone,
            "email"      : email,
        })

    upsert_members(society_id, members)
    return len(members)


# ══════════════════════════════════════════════════════════════
#  SOCIETY MEMBERS DIRECTORY
# ══════════════════════════════════════════════════════════════

@app.route("/members")
@society_required
def members_directory():
    society_id = session["society_id"]
    members    = get_members(society_id)
    return render_template("members.html",
                           society_name=session["society_name"],
                           members=members)


@app.route("/members/upload", methods=["POST"])
@login_required
def members_upload():
    """Upload an Excel file with member directory. Works for both society users and admin."""
    if "excel" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file = request.files["excel"]

    # Society user: use their own society_id from session
    # Admin: society_id must be passed in form
    if session.get("society_id"):
        society_id = session["society_id"]
    else:
        sid = request.form.get("society_id") or request.args.get("society_id")
        if not sid:
            return jsonify({"success": False, "error": "society_id required for admin upload"}), 400
        try:
            society_id = int(sid)
        except (ValueError, TypeError):
            return jsonify({"success": False, "error": "Invalid society_id"}), 400

    try:
        count = _process_member_excel(file, society_id)
        return jsonify({"success": True, "count": count})
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/members/delete-all", methods=["POST"])
@society_required
def members_delete_all():
    delete_all_members(session["society_id"])
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════════════
#  WHATSAPP NOTICE SENDING
# ══════════════════════════════════════════════════════════════

WABA_TOKEN    = os.environ.get("WHATSAPP_TOKEN", "")
WABA_PHONE_ID = os.environ.get("WHATSAPP_PHONE_ID", "")


def _docx_to_pdf(docx_path):
    """Convert a docx file to PDF using LibreOffice. Returns pdf path or None."""
    soffice = get_libreoffice_path()
    if not soffice:
        return None
    out_dir = os.path.dirname(docx_path)
    result  = subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", out_dir, docx_path],
        capture_output=True, timeout=30
    )
    base    = os.path.splitext(os.path.basename(docx_path))[0]
    pdf_path = os.path.join(out_dir, base + ".pdf")
    return pdf_path if os.path.exists(pdf_path) else None


def _send_whatsapp_document(to_phone, pdf_path, caption, filename):
    """
    Send a PDF document via WhatsApp Business Cloud API.
    Returns (success:bool, message:str).
    """
    if not WABA_TOKEN or not WABA_PHONE_ID:
        return False, "WhatsApp credentials not configured. Set WHATSAPP_TOKEN and WHATSAPP_PHONE_ID in Render env vars."

    # Step 1: Upload media to WhatsApp
    with open(pdf_path, "rb") as f:
        upload_resp = http_requests.post(
            f"https://graph.facebook.com/v19.0/{WABA_PHONE_ID}/media",
            headers={"Authorization": f"Bearer {WABA_TOKEN}"},
            files={"file": (filename, f, "application/pdf")},
            data={"messaging_product": "whatsapp"},
            timeout=30
        )
    if upload_resp.status_code != 200:
        return False, f"Media upload failed: {upload_resp.text}"

    media_id = upload_resp.json().get("id")
    if not media_id:
        return False, "No media_id returned from WhatsApp upload"

    # Normalise phone: must be E.164 without +
    phone = to_phone.strip().lstrip("+").replace(" ", "").replace("-", "")
    if len(phone) == 10:
        phone = "91" + phone   # India default

    # Step 2: Send document message
    send_resp = http_requests.post(
        f"https://graph.facebook.com/v19.0/{WABA_PHONE_ID}/messages",
        headers={
            "Authorization" : f"Bearer {WABA_TOKEN}",
            "Content-Type"  : "application/json",
        },
        json={
            "messaging_product": "whatsapp",
            "to"               : phone,
            "type"             : "document",
            "document"         : {
                "id"      : media_id,
                "caption" : caption,
                "filename": filename,
            },
        },
        timeout=30
    )
    if send_resp.status_code == 200:
        return True, "Sent successfully"
    return False, f"Send failed: {send_resp.text}"


@app.route("/whatsapp/preview", methods=["POST"])
@society_required
def whatsapp_preview():
    """
    Given a sess_id + filename (the generated notice docx),
    look up the member phone from the members directory using flat_no.
    Returns member info for the confirmation modal.
    """
    sess_id    = request.json.get("sess_id", "")
    filename   = request.json.get("filename", "")
    flat_combo = request.json.get("flat_combo", "").strip().upper()
    society_id = session["society_id"]

    if not flat_combo:
        return jsonify({"success": False, "error": "Flat number not provided"}), 400

    member = get_member_by_flat(society_id, flat_combo)
    if not member:
        return jsonify({
            "success"  : False,
            "error"    : f"No member found for flat '{flat_combo}' in directory. "
                         "Please upload the member directory first.",
            "not_found": True,
        }), 404

    return jsonify({
        "success"   : True,
        "name"      : member["name"],
        "phone"     : member["phone"],
        "flat_combo": member["flat_combo"],
        "email"     : member.get("email", ""),
        "sess_id"   : sess_id,
        "filename"  : filename,
    })


@app.route("/whatsapp/send", methods=["POST"])
@society_required
def whatsapp_send():
    """
    Convert the saved docx to PDF and send via WhatsApp to the confirmed phone number.
    """
    data      = request.json
    sess_id   = data.get("sess_id", "")
    filename  = data.get("filename", "")
    phone     = data.get("phone", "").strip()
    member_name = data.get("name", "Member")
    society_name = session.get("society_name", "Society")

    docx_path = os.path.join(TEMP_DIR, sess_id, filename)
    if not os.path.exists(docx_path):
        return jsonify({"success": False, "error": "Notice file not found. Please regenerate."}), 404

    # Convert to PDF
    pdf_path = _docx_to_pdf(docx_path)
    if not pdf_path:
        return jsonify({"success": False,
                        "error": "PDF conversion failed — LibreOffice not available on this server."}), 500

    pdf_filename = os.path.splitext(filename)[0] + ".pdf"
    caption=(
    f"Dear {member_name},\n\n"
    f"Please find attached the official notice from {society_name}.\n\n"
    f"Regards,\n{society_name}"
)

    success, msg = _send_whatsapp_document(phone, pdf_path, caption, pdf_filename)
    if success:
        return jsonify({"success": True, "message": f"Notice sent to {phone}"})
    return jsonify({"success": False, "error": msg}), 500


# ── Password Management ────────────────────────────────────────
@app.route("/change-password", methods=["GET","POST"])
@login_required
def change_password():
    error = ""; success = ""
    if request.method == "POST":
        current  = request.form.get("current_password","").strip()
        new_pw   = request.form.get("new_password","").strip()
        confirm  = request.form.get("confirm_password","").strip()
        if new_pw != confirm:
            error = "New passwords do not match."
        elif len(new_pw) < 6:
            error = "Password must be at least 6 characters."
        else:
            sid = session.get("society_id")
            if sid:
                from database import get_society_by_username
                soc = get_society_by_username(session.get("society_username",""))
                # Re-fetch society to verify current password
                from database import get_db
                conn = get_db(); cur = conn.cursor()
                cur.execute("SELECT password FROM societies WHERE id=%s", (sid,))
                row = cur.fetchone(); cur.close(); conn.close()
                if not row or not check_password(current, row["password"]):
                    error = "Current password is incorrect."
                else:
                    change_society_password(sid, new_pw)
                    success = "Password changed successfully!"
            else:
                error = "Admin password must be changed via Render environment variables."
    return render_template("change_password.html", society_name=session.get("society_name",""),
                           error=error, success=success)

@app.route("/admin/change-password", methods=["GET","POST"])
@admin_required
def admin_change_password():
    error = ""; success = ""
    if request.method == "POST":
        current  = request.form.get("current_password","").strip()
        new_pw   = request.form.get("new_password","").strip()
        confirm  = request.form.get("confirm_password","").strip()
        if current != ADMIN_PASSWORD:
            error = "Current admin password is incorrect."
        elif new_pw != confirm:
            error = "New passwords do not match."
        elif len(new_pw) < 6:
            error = "Password must be at least 6 characters."
        else:
            success = "To permanently change admin password, update ADMIN_PASSWORD in Render environment variables."
    return render_template("change_password.html", society_name="Admin",
                           error=error, success=success, is_admin=True)

@app.route("/admin/reset-password/<int:society_id>", methods=["POST"])
@admin_required
def admin_reset_password(society_id):
    import random, string
    new_pw = "".join(random.choices(string.digits, k=6))
    reset_society_password(society_id, new_pw)
    return jsonify({"success": True, "new_password": new_pw})

# ── Monthly Billing ────────────────────────────────────────────
@app.route("/admin/billing")
@admin_required
def admin_billing():
    from database import get_all_bills
    societies = get_all_societies()
    bills     = get_all_bills()
    return render_template("admin_billing.html", societies=societies, bills=bills,
                           society_name="Admin")

@app.route("/admin/billing/create", methods=["POST"])
@admin_required
def admin_create_bill():
    sid     = int(request.form.get("society_id"))
    month   = request.form.get("bill_month")
    year    = request.form.get("bill_year")
    amount  = float(request.form.get("amount", 0))
    desc    = request.form.get("description","")
    create_monthly_bill(sid, month, year, amount, desc)
    return redirect(url_for("admin_billing"))

@app.route("/admin/billing/status/<int:bill_id>", methods=["POST"])
@admin_required
def admin_update_bill_status(bill_id):
    status = request.json.get("status","Unpaid")
    update_bill_status(bill_id, status)
    return jsonify({"success": True})

@app.route("/admin/billing/delete/<int:bill_id>", methods=["POST"])
@admin_required
def admin_delete_bill(bill_id):
    delete_bill(bill_id)
    return jsonify({"success": True})

@app.route("/my-bills")
@society_required
def my_bills():
    bills = get_bills_for_society(session["society_id"])
    return render_template("my_bills.html", bills=bills, society_name=session["society_name"])


@app.route("/my-bills/download/<int:bill_id>")
@society_required
def download_bill_pdf(bill_id):
    """Generate a luxury PDF bill using ReportLab with canvas (no Unicode issues)."""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import textwrap

    bill = get_bill_by_id(bill_id)
    if not bill or bill["society_id"] != session["society_id"]:
        return "Bill not found", 404

    # ── Values ───────────────────────────────────────────────
    amount_fmt  = "Rs. {:,.2f}".format(float(bill["amount"]))
    created_str = bill["created_at"].strftime("%d %B %Y") if bill["created_at"] else "N/A"
    bill_no     = "SNP/{}/{}".format(bill["bill_year"], str(bill["id"]).zfill(4))
    is_paid     = bill["status"] == "Paid"
    period      = "{} {}".format(bill["bill_month"], bill["bill_year"])
    soc_name    = bill["society_name"] or ""
    soc_addr    = bill["society_address"] or "Maharashtra, India"
    regd_no     = bill.get("regd_no") or ""
    desc_text   = bill["description"] or "Monthly platform access fee"

    # ── Colours ──────────────────────────────────────────────
    TEAL      = colors.HexColor("#0ABFA3")
    TEAL2     = colors.HexColor("#069A82")
    TEAL_LITE = colors.HexColor("#E8FAF7")
    TEAL_BDR  = colors.HexColor("#C2EDE6")
    BG        = colors.HexColor("#EEF9F7")
    INK       = colors.HexColor("#063D36")
    INK2      = colors.HexColor("#0A4F48")
    INK3      = colors.HexColor("#4A8A84")
    MUTED     = colors.HexColor("#7ABFB8")
    WHITE     = colors.white
    STATUS_FG = colors.HexColor("#065F46") if is_paid else colors.HexColor("#92400E")
    STATUS_BG = colors.HexColor("#D1FAE5") if is_paid else colors.HexColor("#FEF3C7")
    STATUS_BD = colors.HexColor("#6EE7B7") if is_paid else colors.HexColor("#FCD34D")

    # ── Canvas setup ─────────────────────────────────────────
    buf = io.BytesIO()
    W, H = A4          # 595.27 x 841.89 pt
    c = rl_canvas.Canvas(buf, pagesize=A4)

    # helpers
    def rgb(col):
        return col.red, col.green, col.blue

    def filled_rect(x, y, w, h, fill_col, stroke_col=None, radius=0):
        c.saveState()
        c.setFillColor(fill_col)
        if stroke_col:
            c.setStrokeColor(stroke_col)
            c.setLineWidth(0.8)
        else:
            c.setLineWidth(0)
        if radius:
            c.roundRect(x, y, w, h, radius, fill=1, stroke=1 if stroke_col else 0)
        else:
            c.rect(x, y, w, h, fill=1, stroke=1 if stroke_col else 0)
        c.restoreState()

    def text_line(txt, x, y, font="Helvetica", size=10, color=INK, align="left", max_width=None):
        c.saveState()
        c.setFont(font, size)
        c.setFillColor(color)
        if max_width:
            # simple truncation
            while c.stringWidth(txt, font, size) > max_width and len(txt) > 4:
                txt = txt[:-2]
        if align == "right":
            c.drawRightString(x, y, txt)
        elif align == "center":
            c.drawCentredString(x, y, txt)
        else:
            c.drawString(x, y, txt)
        c.restoreState()

    def wrap_text(txt, x, y, font, size, color, max_width, line_height):
        """Draw wrapped text, return final y position."""
        c.saveState()
        c.setFont(font, size)
        c.setFillColor(color)
        words = txt.split()
        line = ""
        cur_y = y
        for word in words:
            test = (line + " " + word).strip()
            if c.stringWidth(test, font, size) <= max_width:
                line = test
            else:
                if line:
                    c.drawString(x, cur_y, line)
                    cur_y -= line_height
                line = word
        if line:
            c.drawString(x, cur_y, line)
            cur_y -= line_height
        c.restoreState()
        return cur_y

    def hr(y, col=TEAL_BDR, thickness=0.5):
        c.saveState()
        c.setStrokeColor(col)
        c.setLineWidth(thickness)
        c.line(LM, y, W - LM, y)
        c.restoreState()

    LM = 22*mm   # left margin
    RM = W - 22*mm  # right edge
    CW = RM - LM    # content width

    # ════════════════════════════════════════════════════════
    # PAGE BACKGROUND
    # ════════════════════════════════════════════════════════
    filled_rect(0, 0, W, H, BG)

    # ════════════════════════════════════════════════════════
    # HEADER BAND
    # ════════════════════════════════════════════════════════
    HDR_H = 52*mm
    HDR_Y = H - HDR_H
    filled_rect(0, HDR_Y, W, HDR_H, TEAL)
    # decorative circles
    c.saveState()
    c.setFillColor(colors.HexColor("#FFFFFF"))
    c.setFillAlpha(0.07)
    c.circle(W - 25*mm, H - 12*mm, 38*mm, fill=1, stroke=0)
    c.circle(20*mm, HDR_Y - 10*mm, 30*mm, fill=1, stroke=0)
    c.restoreState()
    # top teal-dark bar
    filled_rect(0, H - 3, W, 3, TEAL2)

    # Brand left
    text_line("SocietyNotice", LM, H - 14*mm, "Helvetica-Bold", 18, WHITE)
    text_line("MANAGEMENT SUITE", LM, H - 20*mm, "Helvetica", 7.5,
              colors.HexColor("#CCF0EC"))

    # Invoice right
    text_line("INVOICE", RM, H - 14*mm, "Helvetica-Bold", 26, WHITE, "right")
    text_line("Bill No: {}".format(bill_no), RM, H - 21*mm, "Helvetica", 9,
              colors.HexColor("#CCF0EC"), "right")

    # Divider in header
    c.saveState()
    c.setStrokeColor(colors.HexColor("#FFFFFF"))
    c.setStrokeAlpha(0.2)
    c.setLineWidth(0.5)
    c.line(LM, H - 27*mm, RM, H - 27*mm)
    c.restoreState()

    # Meta row — Billing Period | Issue Date | Amount
    m1x = LM
    m2x = LM + CW * 0.38
    m3x = RM

    meta_label_y = H - 33*mm
    meta_val_y   = H - 39*mm

    text_line("Billing Period", m1x, meta_label_y, "Helvetica", 8,
              colors.HexColor("#99DDD8"))
    text_line(period, m1x, meta_val_y, "Helvetica-Bold", 12, WHITE)

    text_line("Issue Date", m2x, meta_label_y, "Helvetica", 8,
              colors.HexColor("#99DDD8"))
    text_line(created_str, m2x, meta_val_y, "Helvetica-Bold", 12, WHITE)

    text_line("Total Amount", m3x, meta_label_y, "Helvetica", 8,
              colors.HexColor("#99DDD8"), "right")
    text_line(amount_fmt, m3x, meta_val_y, "Helvetica-Bold", 20, WHITE, "right")

    # ════════════════════════════════════════════════════════
    # BODY  — start just below header
    # ════════════════════════════════════════════════════════
    cy = HDR_Y - 10*mm   # current y cursor (moves downward)

    # ── AMOUNT HERO ──────────────────────────────────────────
    HERO_H = 16*mm
    filled_rect(LM, cy - HERO_H, CW, HERO_H, TEAL_LITE, TEAL_BDR, 6)
    text_line("TOTAL AMOUNT", LM + 5*mm, cy - 7*mm, "Helvetica", 8, MUTED)
    text_line(amount_fmt, RM - 5*mm, cy - 8.5*mm,
              "Helvetica-Bold", 18, INK, "right")
    cy -= HERO_H + 7*mm

    # ── BILLED TO + STATUS (side by side) ────────────────────
    BOX_H  = 26*mm
    LEFT_W = CW * 0.57
    GAP    = 3*mm
    RIGHT_W = CW - LEFT_W - GAP

    # Billed-to box
    filled_rect(LM, cy - BOX_H, LEFT_W, BOX_H, TEAL_LITE, TEAL_BDR, 6)
    text_line("BILLED TO", LM + 4*mm, cy - 6*mm, "Helvetica-Bold", 7.5, MUTED)
    text_line(soc_name, LM + 4*mm, cy - 11.5*mm, "Helvetica-Bold", 11, INK,
              max_width=LEFT_W - 8*mm)
    wrap_text(soc_addr, LM + 4*mm, cy - 16*mm, "Helvetica", 9, INK3,
              LEFT_W - 8*mm, 4*mm)
    if regd_no:
        text_line("Regd. No: {}".format(regd_no), LM + 4*mm, cy - 22*mm,
                  "Helvetica", 8, MUTED, max_width=LEFT_W - 8*mm)

    # Status box
    SX = LM + LEFT_W + GAP
    filled_rect(SX, cy - BOX_H, RIGHT_W, BOX_H, STATUS_BG, STATUS_BD, 6)
    text_line("PAYMENT STATUS", SX + 4*mm, cy - 6*mm,
              "Helvetica-Bold", 7.5, MUTED)
    text_line(bill["status"].upper(), SX + 4*mm, cy - 12.5*mm,
              "Helvetica-Bold", 14, STATUS_FG)
    text_line("Bill generated on {}".format(created_str),
              SX + 4*mm, cy - 18*mm, "Helvetica", 8, MUTED,
              max_width=RIGHT_W - 8*mm)

    cy -= BOX_H + 8*mm

    # ── BILL DETAILS SECTION ─────────────────────────────────
    hr(cy, TEAL_BDR, 0.5)
    cy -= 5*mm
    text_line("BILL DETAILS", LM, cy, "Helvetica-Bold", 8, MUTED)
    cy -= 6*mm

    # Table header
    TH_H = 10*mm
    filled_rect(LM, cy - TH_H, CW, TH_H, TEAL_LITE)
    c.saveState()
    c.setStrokeColor(TEAL)
    c.setLineWidth(1.5)
    c.line(LM, cy - TH_H, LM + CW, cy - TH_H)
    c.line(LM, cy,         LM + CW, cy)
    c.restoreState()

    COL1 = LM + 4*mm
    COL2 = LM + CW * 0.58
    COL3 = RM - 4*mm
    THY  = cy - 6.5*mm

    text_line("DESCRIPTION",   COL1, THY, "Helvetica-Bold", 8, MUTED)
    text_line("PERIOD",        COL2, THY, "Helvetica-Bold", 8, MUTED, "center")
    text_line("AMOUNT",        COL3, THY, "Helvetica-Bold", 8, MUTED, "right")
    cy -= TH_H

    # Table row
    ROW_H = 20*mm
    c.saveState()
    c.setStrokeColor(TEAL_BDR)
    c.setLineWidth(0.5)
    c.line(LM, cy - ROW_H, LM + CW, cy - ROW_H)
    c.restoreState()

    text_line("SocietyNotice Platform - Monthly Subscription",
              COL1, cy - 6*mm, "Helvetica-Bold", 10, INK, max_width=CW*0.54)
    text_line(desc_text, COL1, cy - 11*mm, "Helvetica", 8.5, MUTED,
              max_width=CW * 0.54)
    text_line(period, COL2, cy - 7.5*mm, "Helvetica", 10, INK3, "center")
    text_line(amount_fmt, COL3, cy - 7.5*mm, "Helvetica-Bold", 11, INK, "right")
    cy -= ROW_H

    # Total row
    TOT_H = 12*mm
    filled_rect(LM, cy - TOT_H, CW, TOT_H, TEAL_LITE)
    c.saveState()
    c.setStrokeColor(TEAL)
    c.setLineWidth(1.5)
    c.line(LM, cy, LM + CW, cy)
    c.line(LM, cy - TOT_H, LM + CW, cy - TOT_H)
    c.restoreState()
    text_line("TOTAL", COL1, cy - 7.5*mm, "Helvetica-Bold", 9, MUTED)
    text_line(amount_fmt, COL3, cy - 7*mm, "Helvetica-Bold", 13, TEAL, "right")
    cy -= TOT_H + 8*mm

    # ── INFO GRID (3 cards) ──────────────────────────────────
    CARD_W = (CW - 4*mm) / 3
    CARD_H = 18*mm
    cards = [
        ("BILL NUMBER",    bill_no,                    INK),
        ("BILLING MONTH",  period,                     INK),
        ("PAYMENT STATUS", bill["status"],             STATUS_FG),
    ]
    for i, (lbl, val, vcol) in enumerate(cards):
        cx_ = LM + i * (CARD_W + 2*mm)
        filled_rect(cx_, cy - CARD_H, CARD_W, CARD_H, TEAL_LITE, TEAL_BDR, 5)
        text_line(lbl, cx_ + 4*mm, cy - 6*mm,  "Helvetica-Bold", 7.5, MUTED)
        text_line(val, cx_ + 4*mm, cy - 12.5*mm, "Helvetica-Bold", 10, vcol,
                  max_width=CARD_W - 8*mm)

    cy -= CARD_H + 8*mm

    # ── SYSTEM NOTE ──────────────────────────────────────────
    NOTE_H = 22*mm
    filled_rect(LM, cy - NOTE_H, CW, NOTE_H, TEAL_LITE, TEAL_BDR, 6)
    note_x  = LM + 5*mm
    note_mw = CW - 10*mm
    note_text = (
        "Note: This is a system-generated bill and does not require any physical "
        "signature or stamp. This document is legally valid without a manual signature "
        "as per the terms of the SocietyNotice platform agreement. "
        "For any queries, please contact your platform administrator."
    )
    # Bold "Note:"
    text_line("Note:", note_x, cy - 6*mm, "Helvetica-Bold", 9, INK2)
    note_body = note_text[5:].strip()
    wrap_text(note_body, note_x + 22, cy - 6*mm, "Helvetica", 9, INK3,
              note_mw - 22, 4.5*mm)

    cy -= NOTE_H + 6*mm

    # ── FOOTER ───────────────────────────────────────────────
    hr(cy, TEAL_BDR, 0.5)
    cy -= 5*mm
    footer = "Generated by SocietyNotice Management Suite  |  {}".format(created_str)
    text_line(footer, W / 2, cy, "Helvetica", 8, MUTED, "center")

    # Bottom teal bar
    filled_rect(0, 0, W, 4, TEAL2)

    # ════════════════════════════════════════════════════════
    c.showPage()
    c.save()

    buf.seek(0)
    fname = "Bill_{}_{}.pdf".format(period.replace(" ", "_"), bill_no.replace("/", "_"))
    resp = make_response(buf.read())
    resp.headers["Content-Type"]        = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp

# ══════════════════════════════════════════════════════════════
#  MEMBER PORTAL
# ══════════════════════════════════════════════════════════════
import hashlib as _hl

def _pin_hash(pin): return _hl.sha256(pin.upper().encode()).hexdigest()

def _default_pin(flat_combo: str, pin_format: str = 'no_hyphen') -> str:
    """Return the plain-text default PIN for a flat based on the society's format setting.
    'no_hyphen' → B01-310 becomes B01310  (recommended for Bxx-NNN style flats)
    'flat_combo' → B01-310 stays  B01-310
    """
    if pin_format == 'no_hyphen':
        return flat_combo.replace('-', '').upper()
    return flat_combo.upper()

def portal_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("member_flat"):
            return redirect(url_for("portal_login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/portal")
def portal_home():
    return redirect(url_for("portal_login"))


@app.route("/portal/login", methods=["GET","POST"])
def portal_login():
    error = ""
    if request.method == "POST":
        portal_code = request.form.get("portal_code","").strip().upper()
        flat_combo  = request.form.get("flat_combo","").strip().upper()
        pin         = request.form.get("pin","").strip()

        society = get_society_by_portal_code(portal_code)
        if not society:
            error = "Invalid society code. Please check and try again."
        else:
            member = get_member_by_flat(society["id"], flat_combo)
            if not member:
                error = "Flat number not found in member directory."
            else:
                ml = get_member_login(society["id"], flat_combo)
                if not ml:
                    # First ever login — set default PIN based on society's configured format
                    pin_fmt    = society.get("default_pin_format") or "no_hyphen"
                    def_pin    = _default_pin(flat_combo, pin_fmt)
                    upsert_member_login(society["id"], flat_combo,
                                        _pin_hash(def_pin), must_change=True)
                    ml = get_member_login(society["id"], flat_combo)

                entered_hash = _pin_hash(pin)
                # Backward-compat: members whose record was seeded with flat_combo
                # (old code) can still log in — we upgrade their hash on the fly.
                old_hash = _pin_hash(flat_combo)  # e.g. hash("B01-310")
                if ml["pin_hash"] == old_hash and entered_hash != old_hash:
                    # User typed new-style PIN (B01310) but record has old-style hash
                    # — check new-style and silently upgrade
                    pin_fmt = society.get("default_pin_format") or "no_hyphen"
                    new_default = _default_pin(flat_combo, pin_fmt)
                    if _pin_hash(pin) == _pin_hash(new_default):
                        # Upgrade stored hash to new default
                        upsert_member_login(society["id"], flat_combo,
                                            _pin_hash(new_default), must_change=True)
                        ml = get_member_login(society["id"], flat_combo)
                        entered_hash = _pin_hash(new_default)

                if ml["pin_hash"] != entered_hash:
                    pin_fmt  = society.get("default_pin_format") or "no_hyphen"
                    eg_flat  = "B01-310"
                    eg_pin   = _default_pin(eg_flat, pin_fmt)
                    error = f"Incorrect PIN. Your default PIN is your flat number without hyphen (e.g. {eg_pin} for flat {eg_flat})."
                else:
                    session["member_flat"]     = flat_combo
                    session["member_name"]     = member["name"]
                    session["member_society_id"] = society["id"]
                    session["member_society"]  = society["name"]
                    session["member_portal_code"] = portal_code
                    touch_member_login(society["id"], flat_combo)
                    if ml["must_change"]:
                        return redirect(url_for("portal_set_pin"))
                    return redirect(url_for("portal_dashboard"))

    return render_template("portal_login.html", error=error)


@app.route("/portal/logout")
def portal_logout():
    for k in ["member_flat","member_name","member_society_id",
              "member_society","member_portal_code"]:
        session.pop(k, None)
    return redirect(url_for("portal_login"))


@app.route("/portal/set-pin", methods=["GET","POST"])
@portal_required
def portal_set_pin():
    error = ""; success = ""
    if request.method == "POST":
        new_pin = request.form.get("new_pin","").strip()
        confirm = request.form.get("confirm_pin","").strip()
        if len(new_pin) < 4:
            error = "PIN must be at least 4 characters."
        elif new_pin != confirm:
            error = "PINs do not match."
        else:
            flat_up      = session["member_flat"].upper()
            blocked_pins = {flat_up, flat_up.replace('-', '')}  # block both B01-310 and B01310
            if new_pin.upper() in blocked_pins:
                error = "PIN cannot be the same as your default flat number."
            else:
                update_member_pin(session["member_society_id"],
                                  session["member_flat"], _pin_hash(new_pin))
                return redirect(url_for("portal_dashboard"))
    return render_template("portal_set_pin.html", error=error,
                           society_name=session["member_society"],
                           member_name=session["member_name"])


@app.route("/portal/dashboard")
@portal_required
def portal_dashboard():
    sid     = session["member_society_id"]
    flat    = session["member_flat"]
    notices = get_member_notices(sid, flat)
    member  = get_member_by_flat(sid, flat)
    anns    = get_member_announcements(sid)
    society = get_all_societies()
    soc     = next((s for s in society if s["id"] == sid), {})
    my_tickets = get_member_tickets(sid, flat)
    open_tickets = len([t for t in my_tickets if t["status"] != "Resolved"])
    return render_template("portal_dashboard.html",
                           member=member, notices=notices,
                           announcements=anns, society=soc,
                           society_name=session["member_society"],
                           flat=flat, member_name=session["member_name"],
                           my_tickets=my_tickets[:3],
                           open_tickets=open_tickets)


@app.route("/portal/chat", methods=["POST"])
@csrf.exempt
@portal_required
def portal_chat():
    """AI chatbot — answers questions using the member's own data + society KB as context."""
    try:
        user_msg = request.json.get("message","").strip()
        if not user_msg:
            return jsonify({"reply": "Please type a message."})

        sid    = session["member_society_id"]
        flat   = session["member_flat"]
        name   = session["member_name"]
        soc    = session["member_society"]

        notices = get_member_notices(sid, flat)
        member  = get_member_by_flat(sid, flat)
        anns    = get_member_announcements(sid)

        # ── Build notice context ──────────────────────────────────────────
        notice_lines = []
        total_pending_amount = 0
        for n in notices[:10]:
            paid_str = f", paid on {n.get('payment_date','')}" if n.get('payment_date') else ""
            notice_lines.append(
                f"- Ref {n['ref_no']}: {n['notice_type']} notice, "
                f"issued {n['issued_date']}, notice amount Rs.{n['amount']}, "
                f"status: {n['payment_status']}{paid_str}"
            )
            if n["payment_status"] == "Pending":
                total_pending_amount += int(n.get("amount", 0))

        ann_lines = [f"- {a['title']}: {a['body'][:120]}" for a in anns[:5]]

        # ── Vector KB semantic search ─────────────────────────────────────
        kb_sections = []
        flat_outstanding_line = None
        try:
            q_embed  = embed_query(user_msg)
            vresults = vector_search(sid, q_embed, top_k=6)
            if vresults:
                kb_sections.append("RELEVANT SOCIETY KNOWLEDGE (semantic search):")
                for r in vresults:
                    if r["similarity"] > 0.3:  # only confident matches
                        label = r["kb_type"].upper()
                        kb_sections.append(f"[{label} — {r['doc_name']}]\n{r['chunk_text']}")
                        # Check if this chunk has flat outstanding info
                        if flat.upper() in r["chunk_text"].upper() and "outstanding" in r["kb_type"].lower():
                            for line in r["chunk_text"].splitlines():
                                if flat.upper() in line.upper():
                                    flat_outstanding_line = line.strip()
                                    break
        except Exception as _ve:
            print(f"[VECTOR SEARCH WARN] {_ve}")

        # Fallback to legacy text KB if vector KB is empty
        if not kb_sections:
            kb_entries = get_knowledge(sid)
            for kb in kb_entries:
                if kb["kb_type"] == "outstanding":
                    for line in kb["content"].splitlines():
                        if flat.upper() in line.upper():
                            flat_outstanding_line = line.strip()
                            break
                    kb_sections.append(f"OUTSTANDING AMOUNTS:\n{kb['content'][:3000]}")
                elif kb["kb_type"] == "rules":
                    kb_sections.append(f"SOCIETY RULES:\n{kb['content'][:3000]}")
                else:
                    kb_sections.append(f"SOCIETY INFO:\n{kb['content'][:2000]}")

        kb_text = "\n\n".join(kb_sections) if kb_sections else "No society knowledge base loaded yet."

        outstanding_summary = (
            f"Outstanding for Flat {flat} (from society records): {flat_outstanding_line}"
            if flat_outstanding_line
            else f"Outstanding notices for Flat {flat}: Rs.{total_pending_amount:,} (based on {len([n for n in notices if n['payment_status']=='Pending'])} pending notice(s))"
        )

        context = f"""You are the official AI assistant for {soc} housing society.
You are speaking with {name}, resident of Flat {flat}.

STRICT RULES:
- Answer ONLY questions related to this member's notices, outstanding dues, society rules, announcements, or general society information.
- If a member asks something unrelated to the society (e.g. news, politics, sports, general knowledge), respond: "I'm here to assist with {soc} society matters only. For other queries, please use a general search engine."
- NEVER say "I don't know" — if you lack the data, say "Please contact the society office for this information."
- Use the KNOWLEDGE BASE data for outstanding amount queries, not just notice amounts.
- For notice-related questions (ref no, due date, notice amount), refer to NOTICES section.
- Be helpful, polite, and concise (2-4 sentences). Reply in the same language the member uses.

MEMBER DATA:
Name: {name}
Flat: {flat}
Society: {soc}
Phone: {member.get('phone','N/A') if member else 'N/A'}

OUTSTANDING SUMMARY:
{outstanding_summary}

NOTICES ISSUED TO THIS MEMBER:
{chr(10).join(notice_lines) if notice_lines else 'No notices issued yet.'}

RECENT SOCIETY ANNOUNCEMENTS:
{chr(10).join(ann_lines) if ann_lines else 'No announcements yet.'}

{kb_text}"""

        reply = call_groq(context, user_msg)
        return jsonify({"reply": reply})

    except Exception as e:
        return jsonify({"reply": f"Sorry, I'm having trouble right now. ({str(e)[:60]})"}), 500


# ── Admin: Portal Code management & Announcements ─────────────
@app.route("/admin/set-portal-code/<int:society_id>", methods=["POST"])
@admin_required
def admin_set_portal_code(society_id):
    code = request.json.get("code","").strip().upper()
    if not code:
        return jsonify({"success": False, "error": "Code cannot be empty"}), 400
    try:
        set_portal_code(society_id, code)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/admin/reset-member-pin/<int:society_id>", methods=["POST"])
@admin_required
def admin_reset_member_pin(society_id):
    flat = request.json.get("flat_combo","").strip().upper()
    if not flat:
        return jsonify({"success": False, "error": "Flat number required"}), 400
    pin_fmt     = get_society_pin_format(society_id)
    default_pin = _default_pin(flat, pin_fmt)
    reset_member_pin(society_id, flat, default_pin)
    return jsonify({"success": True, "default_pin": default_pin})


@app.route("/portal/announcements/create", methods=["POST"])
@login_required
def create_announcement_route():
    sid = session.get("society_id")
    if not sid:
        return jsonify({"success": False, "error": "Not a society user"}), 403
    data = request.json
    create_announcement(sid, data.get("title",""), data.get("body",""),
                        data.get("category","General"))
    return jsonify({"success": True})


@app.route("/portal/announcements/delete/<int:ann_id>", methods=["POST"])
@login_required
def delete_announcement_route(ann_id):
    delete_announcement(ann_id)
    return jsonify({"success": True})


@app.route("/portal/announcements")
@login_required
def manage_announcements():
    sid  = session.get("society_id")
    anns = get_member_announcements(sid) if sid else []
    return render_template("announcements.html",
                           announcements=anns,
                           society_name=session.get("society_name",""))

# ══════════════════════════════════════════════════════════════
#  KNOWLEDGE BASE — society can upload outstanding Excel + rules text
# ══════════════════════════════════════════════════════════════

@app.route("/knowledge")
@society_required
def knowledge_page():
    sid  = session["society_id"]
    kb   = {r["kb_type"]: r for r in get_knowledge(sid)}
    docs = get_kb_documents(sid)
    total_chunks = get_kb_chunk_count(sid)
    # Group docs by kb_type for easy template access
    docs_by_type = {}
    for d in docs:
        docs_by_type.setdefault(d["kb_type"], []).append(d)
    return render_template("knowledge.html",
                           society_name=session["society_name"],
                           kb=kb,
                           docs_by_type=docs_by_type,
                           total_chunks=total_chunks)


@app.route("/knowledge/upload-doc", methods=["POST"])
@society_required
def knowledge_upload_doc():
    """
    Upload a PDF/DOCX/Excel/TXT document to the vector knowledge base.
    Extracts text, chunks it, embeds via Groq, stores in pgvector.
    """
    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    file    = request.files["file"]
    kb_type = request.form.get("kb_type", "general")
    if file.filename == "":
        return jsonify({"success": False, "error": "No file selected"}), 400

    allowed = {"pdf", "docx", "doc", "xlsx", "xls", "txt"}
    ext     = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in allowed:
        return jsonify({"success": False, "error": f"Unsupported file type .{ext}. Use PDF, DOCX, Excel or TXT."}), 400

    try:
        file_bytes = file.read()
        doc_name   = file.filename
        society_id = session["society_id"]

        # Full pipeline: extract → chunk → embed
        chunks, embeddings = process_document(file_bytes, doc_name)

        # Pair and store in pgvector
        pairs = list(zip(chunks, embeddings))
        save_kb_chunks(society_id, kb_type, doc_name, ext, pairs)

        return jsonify({
            "success": True,
            "doc_name": doc_name,
            "chunks": len(chunks),
            "kb_type": kb_type,
            "message": f"✅ {doc_name} uploaded — {len(chunks)} chunks indexed in vector KB"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/knowledge/delete-doc", methods=["POST"])
@society_required
def knowledge_delete_doc():
    """Delete a document and all its chunks from the vector KB."""
    data     = request.json
    doc_name = data.get("doc_name", "")
    kb_type  = data.get("kb_type", "")
    if not doc_name or not kb_type:
        return jsonify({"success": False, "error": "doc_name and kb_type required"}), 400
    try:
        delete_kb_document(session["society_id"], doc_name, kb_type)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/knowledge/search", methods=["POST"])
@society_required
def knowledge_search():
    """Test semantic search endpoint — for admin to verify the KB is working."""
    query = request.json.get("query", "").strip()
    if not query:
        return jsonify({"success": False, "error": "Query required"}), 400
    try:
        q_embed = embed_query(query)
        results = vector_search(session["society_id"], q_embed, top_k=5)
        return jsonify({"success": True, "results": results})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/knowledge/upload-outstanding", methods=["POST"])
@society_required
def knowledge_upload_outstanding():
    """Parse an Excel of outstanding amounts and store as text in the KB."""
    if "excel" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400
    file = request.files["excel"]
    try:
        df = pd.read_excel(file)
        df.columns = [str(c).strip() for c in df.columns]

        # Detect flat, amount, and member columns by matching common header names
        flat_col   = None
        amount_col = None
        member_col = None
        for col in df.columns:
            cl = col.lower().replace(' ', '').replace('_', '').replace('.', '')
            if flat_col is None and any(k in cl for k in ('flat', 'unit', 'aptno', 'flatno')):
                flat_col = col
            elif amount_col is None and any(k in cl for k in ('outstanding', 'amount', 'dues', 'balance', 'pending', 'total', 'owed', 'arrear')):
                amount_col = col
            elif member_col is None and any(k in cl for k in ('member', 'name', 'owner', 'resident')):
                member_col = col

        lines = ["Flat | Outstanding Amount | Member"]
        lines.append("-" * 50)

        if flat_col and amount_col:
            # Structured extraction using detected columns
            for _, row in df.iterrows():
                flat   = str(row[flat_col]).strip()
                amt    = str(row[amount_col]).strip()
                member = str(row[member_col]).strip() if member_col else ''
                if not flat or flat.lower() == 'nan' or amt.lower() == 'nan':
                    continue
                # Skip header-like rows
                if flat.lower() in ('flat', 'flat no', 'flat_no', 'unit'):
                    continue
                member = '' if member.lower() == 'nan' else member
                lines.append(f"{flat} | {amt} | {member}")
        else:
            # Fallback: assume first col = flat, last numeric col = amount
            for _, row in df.iterrows():
                vals = [str(v).strip() for v in row.values]
                flat = vals[0] if vals else ''
                if not flat or flat.lower() in ('nan', 'flat', 'flat no'):
                    continue
                # Find rightmost numeric value as amount
                amt = ''
                member = ''
                for v in reversed(vals[1:]):
                    import re as _re2
                    if _re2.sub(r'[^\d.]', '', v):
                        amt = v
                        break
                lines.append(f"{flat} | {amt} | {member}")

        content = "\n".join(lines)
        upsert_knowledge(session["society_id"], "outstanding", content)
        return jsonify({"success": True, "rows": len(df)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/knowledge/save-rules", methods=["POST"])
@society_required
def knowledge_save_rules():
    """Save society rules/regulations text to the KB."""
    content = request.json.get("content", "").strip()
    if not content:
        return jsonify({"success": False, "error": "Content cannot be empty"}), 400
    upsert_knowledge(session["society_id"], "rules", content)
    return jsonify({"success": True})


@app.route("/knowledge/save-general", methods=["POST"])
@society_required
def knowledge_save_general():
    """Save any general society info text to the KB."""
    content = request.json.get("content", "").strip()
    if not content:
        return jsonify({"success": False, "error": "Content cannot be empty"}), 400
    upsert_knowledge(session["society_id"], "general", content)
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════════════
#  MEMBER PORTAL — Change PIN
# ══════════════════════════════════════════════════════════════

@app.route("/portal/change-pin", methods=["GET", "POST"])
@portal_required
def portal_change_pin():
    error = ""; success = ""
    if request.method == "POST":
        current_pin = request.form.get("current_pin", "").strip()
        new_pin     = request.form.get("new_pin", "").strip()
        confirm_pin = request.form.get("confirm_pin", "").strip()
        sid  = session["member_society_id"]
        flat = session["member_flat"]
        ml   = get_member_login(sid, flat)
        if not ml or ml["pin_hash"] != _pin_hash(current_pin):
            error = "Current PIN is incorrect."
        elif len(new_pin) < 4:
            error = "New PIN must be at least 4 characters."
        elif new_pin != confirm_pin:
            error = "New PINs do not match."
        else:
            flat_up      = flat.upper()
            blocked_pins = {flat_up, flat_up.replace('-', '')}
            if new_pin.upper() in blocked_pins:
                error = "PIN cannot be the same as your flat number."
            else:
                update_member_pin(sid, flat, _pin_hash(new_pin))
                success = "PIN changed successfully!"
    return render_template("portal_change_pin.html",
                           society_name=session["member_society"],
                           member_name=session["member_name"],
                           flat=session["member_flat"],
                           error=error, success=success)


# ══════════════════════════════════════════════════════════════
#  SOCIETY (Committee) — Reset member PIN
# ══════════════════════════════════════════════════════════════

@app.route("/society/reset-member-pin", methods=["POST"])
@society_required
def society_reset_member_pin():
    """Allows committee (society login) to reset any member's PIN to the default."""
    flat = request.json.get("flat_combo", "").strip().upper()
    if not flat:
        return jsonify({"success": False, "error": "Flat number required"}), 400
    sid         = session["society_id"]
    pin_fmt     = get_society_pin_format(sid)
    default_pin = _default_pin(flat, pin_fmt)
    reset_member_pin(sid, flat, default_pin)
    return jsonify({"success": True, "default_pin": default_pin, "flat": flat})


# ══════════════════════════════════════════════════════════════
#  MEMBER TICKETS / COMPLAINTS
# ══════════════════════════════════════════════════════════════

TICKET_CATEGORIES = [
    "Maintenance", "Water / Plumbing", "Electricity", "Lift / Elevator",
    "Parking", "Security", "Cleanliness / Housekeeping", "Noise Complaint",
    "Neighbour Dispute", "Common Area", "Administrative", "Other"
]

@app.route("/portal/tickets")
@portal_required
def portal_tickets():
    sid    = session["member_society_id"]
    flat   = session["member_flat"]
    tickets = get_member_tickets(sid, flat)
    return render_template("portal_tickets.html",
                           tickets=tickets,
                           categories=TICKET_CATEGORIES,
                           society_name=session["member_society"],
                           member_name=session["member_name"],
                           flat=flat)


@app.route("/portal/tickets/create", methods=["POST"])
@csrf.exempt
@portal_required
def portal_create_ticket():
    sid  = session["member_society_id"]
    flat = session["member_flat"]
    name = session["member_name"]
    data = request.json or {}
    category    = data.get("category", "General").strip()
    subject     = data.get("subject", "").strip()
    description = data.get("description", "").strip()
    priority    = data.get("priority", "Normal").strip()
    if not subject or not description:
        return jsonify({"success": False, "error": "Subject and description are required"}), 400
    tid = create_ticket(sid, flat, name, category, subject, description, priority)
    return jsonify({"success": True, "ticket_id": tid})


# ── Society (committee) ticket management ──────────────────────

@app.route("/tickets")
@society_required
def society_tickets():
    sid    = session["society_id"]
    status = request.args.get("status", "")
    tickets = get_all_tickets(sid, status or None)
    counts = {
        "all":        len(get_all_tickets(sid)),
        "open":       len(get_all_tickets(sid, "Open")),
        "inprogress": len(get_all_tickets(sid, "In Progress")),
        "resolved":   len(get_all_tickets(sid, "Resolved")),
    }
    return render_template("society_tickets.html",
                           tickets=tickets, counts=counts,
                           selected_status=status,
                           society_name=session["society_name"])


@app.route("/tickets/update/<int:ticket_id>", methods=["POST"])
@society_required
def society_update_ticket(ticket_id):
    data   = request.json or {}
    status = data.get("status", "Open")
    note   = data.get("committee_note", "").strip()
    update_ticket_status(ticket_id, status, note)
    log_audit(session["society_id"], session["society_name"],
              "TICKET_UPDATE", f"Ticket #{ticket_id} → {status}")
    return jsonify({"success": True})


# ══════════════════════════════════════════════════════════════
#  PHASE 2: RAZORPAY PAYMENT INTEGRATION
# ══════════════════════════════════════════════════════════════

def _razorpay_client():
    if not RAZORPAY_KEY_ID or not RAZORPAY_KEY_SECRET:
        return None
    import razorpay
    return razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))


@app.route("/portal/pay/create-order", methods=["POST"])
@portal_required
def portal_create_payment_order():
    """Create a Razorpay order for a pending notice."""
    client = _razorpay_client()
    if not client:
        return jsonify({"success": False,
                        "error": "Payment gateway not configured. Contact society office."}), 503

    sid       = session["member_society_id"]
    flat      = session["member_flat"]
    name      = session["member_name"]
    data      = request.json or {}
    notice_id = data.get("notice_id")
    amount    = data.get("amount")          # in rupees

    if not notice_id or not amount:
        return jsonify({"success": False, "error": "notice_id and amount required"}), 400

    try:
        amount_paise = int(float(amount) * 100)
        order = client.order.create({
            "amount":   amount_paise,
            "currency": "INR",
            "receipt":  f"notice_{notice_id}_{flat}",
            "notes":    {"flat": flat, "society_id": str(sid), "notice_id": str(notice_id)},
        })
        create_payment_order(sid, flat, name, notice_id, order["id"], amount,
                             f"Maintenance notice payment")
        return jsonify({
            "success":  True,
            "order_id": order["id"],
            "key_id":   RAZORPAY_KEY_ID,
            "amount":   amount_paise,
            "currency": "INR",
            "name":     session["member_society"],
            "member":   name,
            "flat":     flat,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/razorpay/webhook", methods=["POST"])
@csrf.exempt
def razorpay_webhook():
    """Razorpay sends payment.captured event here. Verify signature, mark paid."""
    payload   = request.get_data()
    signature = request.headers.get("X-Razorpay-Signature", "")

    if RAZORPAY_KEY_SECRET:
        expected = hmac.new(
            RAZORPAY_KEY_SECRET.encode(),
            payload,
            _hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(expected, signature):
            return jsonify({"error": "Invalid signature"}), 400

    try:
        event = json.loads(payload)
        if event.get("event") == "payment.captured":
            payment = event["payload"]["payment"]["entity"]
            order_id   = payment.get("order_id")
            payment_id = payment.get("id")
            row = confirm_payment(order_id, payment_id)
            if row and row.get("notice_id"):
                update_payment(row["notice_id"], "Paid",
                               datetime.now().strftime("%d-%m-%Y"),
                               int(payment.get("amount", 0) / 100), "Paid via Razorpay")
                log_audit(row["society_id"], row["flat_combo"],
                          "PAYMENT", f"Razorpay {payment_id} Rs.{row['amount']}")
                # Send email receipt
                _send_payment_receipt_email(row)
    except Exception as e:
        print(f"[WEBHOOK ERROR] {e}")

    return jsonify({"status": "ok"})


@app.route("/portal/pay/verify", methods=["POST"])
@portal_required
def portal_verify_payment():
    """Called by frontend after Razorpay checkout completes (client-side confirmation)."""
    data       = request.json or {}
    order_id   = data.get("razorpay_order_id", "")
    payment_id = data.get("razorpay_payment_id", "")
    signature  = data.get("razorpay_signature", "")

    if RAZORPAY_KEY_SECRET:
        body     = f"{order_id}|{payment_id}"
        expected = hmac.new(
            RAZORPAY_KEY_SECRET.encode(), body.encode(), _hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(expected, signature):
            return jsonify({"success": False, "error": "Signature mismatch"}), 400

    row = confirm_payment(order_id, payment_id)
    if row and row.get("notice_id"):
        update_payment(row["notice_id"], "Paid",
                       datetime.now().strftime("%d-%m-%Y"),
                       int(float(data.get("amount", 0)) / 100), "Paid via Razorpay")
        log_audit(session["member_society_id"], session["member_flat"],
                  "PAYMENT", f"Verified Rs.{row['amount']}")
        _send_payment_receipt_email(row)
        return jsonify({"success": True, "message": "Payment confirmed!"})

    return jsonify({"success": False, "error": "Could not confirm payment"}), 500


# ══════════════════════════════════════════════════════════════
#  PHASE 3: EMAIL DELIVERY RECEIPTS (SendGrid)
# ══════════════════════════════════════════════════════════════

def _send_email(to_email: str, subject: str, html_body: str) -> bool:
    """Send an email via SendGrid. Returns True on success."""
    if not SENDGRID_API_KEY or not to_email:
        return False
    try:
        payload = {
            "personalizations": [{"to": [{"email": to_email}]}],
            "from": {"email": SENDGRID_FROM_EMAIL, "name": "SocietyNotice"},
            "subject": subject,
            "content": [{"type": "text/html", "value": html_body}],
        }
        resp = http_requests.post(
            "https://api.sendgrid.com/v3/mail/send",
            headers={"Authorization": f"Bearer {SENDGRID_API_KEY}",
                     "Content-Type": "application/json"},
            json=payload, timeout=10
        )
        return resp.status_code in (200, 202)
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False


def _send_payment_receipt_email(payment_row: dict):
    sid   = payment_row.get("society_id")
    flat  = payment_row.get("flat_combo", "")
    member = get_member_by_flat(sid, flat)
    if not member or not member.get("email"):
        return
    socs = get_all_societies()
    soc  = next((s for s in socs if s["id"] == sid), {})
    html = f"""
    <div style="font-family:sans-serif;max-width:600px;margin:auto;padding:32px;">
      <h2 style="color:#0ABFA3;">✅ Payment Confirmed</h2>
      <p>Dear <strong>{member['name']}</strong>,</p>
      <p>Your maintenance payment of <strong>₹{payment_row['amount']:,.0f}</strong>
         has been received for Flat <strong>{flat}</strong>.</p>
      <table style="width:100%;border-collapse:collapse;margin:20px 0;">
        <tr><td style="padding:8px;color:#888;">Society</td>
            <td style="padding:8px;font-weight:600;">{soc.get('name','')}</td></tr>
        <tr style="background:#f9f9f9;">
          <td style="padding:8px;color:#888;">Date</td>
          <td style="padding:8px;">{datetime.now().strftime('%d %b %Y %I:%M %p')}</td></tr>
        <tr><td style="padding:8px;color:#888;">Amount</td>
            <td style="padding:8px;font-weight:600;color:#0ABFA3;">₹{payment_row['amount']:,.0f}</td></tr>
      </table>
      <p style="color:#888;font-size:12px;">This is an auto-generated receipt from SocietyNotice.</p>
    </div>"""
    _send_email(member["email"],
                f"Payment Receipt — {soc.get('name', 'Society')} — Flat {flat}",
                html)


def _send_notice_delivery_email(member_email: str, member_name: str,
                                flat: str, society_name: str,
                                ref_no: str, amount: int, issued_date: str):
    """Email sent when a notice is generated for a member."""
    if not member_email:
        return False
    html = f"""
    <div style="font-family:sans-serif;max-width:600px;margin:auto;padding:32px;border:1px solid #e5e5e5;border-radius:12px;">
      <div style="background:#0ABFA3;padding:20px;border-radius:8px 8px 0 0;margin:-32px -32px 24px;">
        <h2 style="color:#fff;margin:0;">{society_name}</h2>
        <p style="color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:13px;">Maintenance Notice</p>
      </div>
      <p>Dear <strong>{member_name}</strong>,</p>
      <p>A maintenance notice has been issued for your flat.</p>
      <table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:14px;">
        <tr style="background:#f4fdfb;"><td style="padding:10px;color:#888;">Flat No</td>
            <td style="padding:10px;font-weight:600;">{flat}</td></tr>
        <tr><td style="padding:10px;color:#888;">Ref No</td>
            <td style="padding:10px;font-family:monospace;">{ref_no}</td></tr>
        <tr style="background:#f4fdfb;"><td style="padding:10px;color:#888;">Amount Due</td>
            <td style="padding:10px;font-weight:700;color:#D97706;">₹{amount:,}</td></tr>
        <tr><td style="padding:10px;color:#888;">Issued Date</td>
            <td style="padding:10px;">{issued_date}</td></tr>
      </table>
      <p>Please log in to the Member Portal to view and pay your notice.</p>
      <p style="color:#888;font-size:12px;">Regards,<br>{society_name} Management</p>
    </div>"""
    return _send_email(member_email,
                       f"Maintenance Notice — {society_name} — Flat {flat}",
                       html)


# ══════════════════════════════════════════════════════════════
#  PHASE 4: PWA — service worker and manifest
# ══════════════════════════════════════════════════════════════

@app.route("/manifest.json")
def pwa_manifest():
    manifest = {
        "name": "SocietyNotice Member Portal",
        "short_name": "SocietyNotice",
        "description": "Your society notices, dues, and information — all in one place",
        "start_url": "/portal/dashboard",
        "display": "standalone",
        "background_color": "#C8F0EA",
        "theme_color": "#0ABFA3",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"},
        ]
    }
    resp = make_response(json.dumps(manifest))
    resp.headers["Content-Type"] = "application/manifest+json"
    return resp


@app.route("/sw.js")
def service_worker():
    sw = """
const CACHE = 'societynotice-v1';
const OFFLINE_URLS = ['/portal/dashboard', '/portal/login'];
self.addEventListener('install', e =>
  e.waitUntil(caches.open(CACHE).then(c => c.addAll(OFFLINE_URLS))));
self.addEventListener('fetch', e => {
  if (e.request.method !== 'GET') return;
  e.respondWith(
    fetch(e.request).catch(() => caches.match(e.request))
  );
});
"""
    resp = make_response(sw)
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


# ══════════════════════════════════════════════════════════════
#  PHASE 6: ANALYTICS DASHBOARD
# ══════════════════════════════════════════════════════════════

@app.route("/analytics")
@society_required
def analytics_dashboard():
    sid  = session["society_id"]
    data = get_analytics(sid)
    return render_template("analytics.html",
                           society_name=session["society_name"],
                           data=data)


# ══════════════════════════════════════════════════════════════
#  PHASE 7: DPDPA COMPLIANCE
# ══════════════════════════════════════════════════════════════

@app.route("/portal/privacy")
def portal_privacy():
    return render_template("portal_privacy.html",
                           society_name=session.get("member_society", "Society"))


@app.route("/portal/consent", methods=["POST"])
def portal_give_consent():
    """Record explicit DPDPA consent from member."""
    sid  = session.get("member_society_id")
    flat = session.get("member_flat")
    if sid and flat:
        record_consent(sid, flat, request.remote_addr)
        session["has_consent"] = True
    return jsonify({"success": True})


@app.route("/portal/data-request", methods=["POST"])
@portal_required
def portal_data_request():
    """DPDPA right to erasure / data deletion request."""
    sid  = session["member_society_id"]
    flat = session["member_flat"]
    action = request.json.get("action", "")
    if action == "delete":
        delete_member_data(sid, flat)
        log_audit(sid, flat, "DATA_DELETION", "Member requested data deletion",
                  request.remote_addr)
        session.clear()
        return jsonify({"success": True,
                        "message": "Your data has been deleted. You have been signed out."})
    return jsonify({"success": False, "error": "Unknown action"}), 400


@app.route("/portal/audit-log")
@portal_required
def portal_my_audit():
    """Member can see their own activity log (DPDPA transparency)."""
    sid  = session["member_society_id"]
    flat = session["member_flat"]
    logs = get_audit_log(sid, limit=50)
    my_logs = [l for l in logs if l.get("actor") == flat]
    return jsonify({"logs": [
        {"action": l["action"], "detail": l.get("detail",""),
         "when": l["created_at"].strftime("%d %b %Y %H:%M") if l.get("created_at") else ""}
        for l in my_logs
    ]})


@app.route("/admin/audit")
@admin_required
def admin_audit_log():
    """Admin view of all audit events."""
    sid  = request.args.get("society_id")
    logs = get_audit_log(int(sid), 200) if sid else []
    societies = get_all_societies()
    return render_template("admin_audit.html",
                           logs=logs, societies=societies,
                           selected_sid=sid,
                           society_name="Admin")


if __name__ == "__main__":
    lo = get_libreoffice_path()
    print("="*55)
    print("✅ Society Notice App → http://localhost:5000")
    print(f"🔒 Admin: admin / {ADMIN_PASSWORD}")
    print(f"📄 LibreOffice: {lo or '❌ NOT FOUND'}")
    print("="*55)
    app.run(debug=True)